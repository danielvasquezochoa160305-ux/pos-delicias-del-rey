import sqlite3
import json
import os
import io
import urllib.request
import urllib.parse
from datetime import datetime
from flask import Flask, request, jsonify, send_from_directory, send_file
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

app = Flask(__name__, static_folder='public')
DB_PATH = os.path.join(os.path.dirname(__file__), 'pos.db')

# ─── DB HELPERS ──────────────────────────────────────
def get_db():
    conn = sqlite3.connect(DB_PATH, timeout=10, check_same_thread=False)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA synchronous=NORMAL")
    conn.execute("PRAGMA cache_size=10000")
    conn.execute("PRAGMA foreign_keys=ON")
    return conn

def init_db():
    with get_db() as conn:
        conn.executescript("""
        CREATE TABLE IF NOT EXISTS products (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            category TEXT NOT NULL DEFAULT 'General',
            price REAL NOT NULL,
            cost REAL NOT NULL DEFAULT 0,
            stock REAL NOT NULL DEFAULT 0,
            unit TEXT NOT NULL DEFAULT 'unidad',
            low_stock_alert REAL NOT NULL DEFAULT 5,
            infinite_stock INTEGER NOT NULL DEFAULT 0,
            active INTEGER NOT NULL DEFAULT 1,
            created_at TEXT NOT NULL DEFAULT (datetime('now','localtime'))
        );
        CREATE TABLE IF NOT EXISTS sales (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            subtotal REAL NOT NULL,
            total REAL NOT NULL,
            payment_method TEXT NOT NULL DEFAULT 'efectivo',
            notes TEXT,
            created_at TEXT NOT NULL DEFAULT (datetime('now','localtime'))
        );
        CREATE TABLE IF NOT EXISTS sale_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            sale_id INTEGER NOT NULL,
            product_id INTEGER,
            product_name TEXT NOT NULL,
            price REAL NOT NULL,
            quantity REAL NOT NULL,
            subtotal REAL NOT NULL,
            FOREIGN KEY (sale_id) REFERENCES sales(id)
        );
        CREATE TABLE IF NOT EXISTS cash_registers (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            status TEXT NOT NULL DEFAULT 'abierta' CHECK(status IN ('abierta','cerrada')),
            opening_balance REAL NOT NULL DEFAULT 0,
            counted_cash REAL,
            total_sales REAL NOT NULL DEFAULT 0,
            total_cash_sales REAL NOT NULL DEFAULT 0,
            total_in REAL NOT NULL DEFAULT 0,
            total_out REAL NOT NULL DEFAULT 0,
            expected_cash REAL,
            difference REAL,
            notes TEXT NOT NULL DEFAULT '',
            opened_at TEXT NOT NULL DEFAULT (datetime('now','localtime')),
            closed_at TEXT
        );
        CREATE TABLE IF NOT EXISTS comandas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            customer_name TEXT NOT NULL DEFAULT '',
            items TEXT NOT NULL,
            notes TEXT NOT NULL DEFAULT '',
            status TEXT NOT NULL DEFAULT 'pendiente' CHECK(status IN ('pendiente','listo','entregado','cancelado')),
            created_at TEXT NOT NULL DEFAULT (datetime('now','localtime'))
        );
        CREATE TABLE IF NOT EXISTS pos_terminals (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            color TEXT NOT NULL DEFAULT '#3b82f6',
            icon TEXT NOT NULL DEFAULT '🛒',
            active INTEGER NOT NULL DEFAULT 1,
            created_at TEXT NOT NULL DEFAULT (datetime('now','localtime'))
        );
        INSERT OR IGNORE INTO pos_terminals (id, name, color, icon) VALUES (1, 'Punto de Venta', '#3b82f6', '🛒');
        CREATE TABLE IF NOT EXISTS settings (
            key TEXT PRIMARY KEY,
            value TEXT NOT NULL DEFAULT ''
        );
        INSERT OR IGNORE INTO settings (key, value) VALUES ('whatsapp_phone', '');
        INSERT OR IGNORE INTO settings (key, value) VALUES ('whatsapp_apikey', '');
        INSERT OR IGNORE INTO settings (key, value) VALUES ('whatsapp_enabled', '0');
        INSERT OR IGNORE INTO settings (key, value) VALUES ('greenapi_instance', '');
        INSERT OR IGNORE INTO settings (key, value) VALUES ('greenapi_token', '');
        INSERT OR IGNORE INTO settings (key, value) VALUES ('greenapi_enabled', '0');
        INSERT OR IGNORE INTO settings (key, value) VALUES ('negocio_nombre', 'Mi Cafetería');
        INSERT OR IGNORE INTO settings (key, value) VALUES ('pin_admin', '1623');
        INSERT OR IGNORE INTO settings (key, value) VALUES ('pin_worker', '0000');
        CREATE TABLE IF NOT EXISTS returns (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            sale_id INTEGER NOT NULL,
            total REAL NOT NULL,
            reason TEXT NOT NULL DEFAULT '',
            payment_method TEXT NOT NULL DEFAULT 'efectivo',
            created_at TEXT NOT NULL DEFAULT (datetime('now','localtime')),
            FOREIGN KEY (sale_id) REFERENCES sales(id)
        );
        CREATE TABLE IF NOT EXISTS return_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            return_id INTEGER NOT NULL,
            product_id INTEGER,
            product_name TEXT NOT NULL,
            price REAL NOT NULL,
            quantity REAL NOT NULL,
            subtotal REAL NOT NULL,
            FOREIGN KEY (return_id) REFERENCES returns(id)
        );
        CREATE TABLE IF NOT EXISTS accounts (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            code TEXT NOT NULL UNIQUE,
            name TEXT NOT NULL,
            type TEXT NOT NULL CHECK(type IN ('ingreso','costo','gasto_operativo','gasto_admin','otro_ingreso','otro_gasto')),
            active INTEGER NOT NULL DEFAULT 1
        );
        CREATE TABLE IF NOT EXISTS journal_entries (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            date TEXT NOT NULL,
            account_id INTEGER NOT NULL,
            description TEXT NOT NULL,
            entry_type TEXT NOT NULL CHECK(entry_type IN ('ingreso','egreso')),
            amount REAL NOT NULL,
            reference TEXT NOT NULL DEFAULT '',
            created_at TEXT NOT NULL DEFAULT (datetime('now','localtime')),
            FOREIGN KEY (account_id) REFERENCES accounts(id)
        );
        INSERT OR IGNORE INTO accounts (id,code,name,type) VALUES
        -- INGRESOS
        (1,'4001','Ventas Punto de Venta','ingreso'),
        (2,'4002','Ventas Institucional','ingreso'),
        -- COSTOS DE VENTAS
        (3,'5001','Compras Punto de Venta','costo'),
        (4,'5002','Cacaos','costo'),
        (5,'5003','Buñuelos','costo'),
        (6,'5004','Gaseosas y Desechables','costo'),
        (7,'5005','Fritos','costo'),
        (8,'5006','Otros Costos','costo'),
        -- GASTOS FIJOS Y VARIABLES
        (9,'6001','Nómina','gasto_operativo'),
        (10,'6002','Arriendo','gasto_operativo'),
        (11,'6003','Servicios Públicos','gasto_operativo'),
        (12,'6004','Internet','gasto_operativo'),
        (13,'6005','Prestaciones Sociales','gasto_operativo'),
        (14,'6006','Mercadeo','gasto_operativo'),
        (15,'6007','Fletes','gasto_operativo'),
        (16,'6008','Fletes Buñuelos','gasto_operativo'),
        (17,'6009','Fumigación','gasto_operativo'),
        (18,'6010','Dotación','gasto_operativo'),
        (19,'6011','Depreciación','gasto_operativo'),
        (20,'6012','Declaración','gasto_operativo'),
        (21,'6013','Licencia POS','gasto_operativo'),
        -- GASTOS FINANCIEROS (gasto_admin)
        (22,'7001','Préstamo','gasto_admin'),
        (23,'7002','4x10000','gasto_admin'),
        -- OTROS GASTOS (códigos 8xxx)
        (24,'8001','Compra de Equipos','otro_gasto'),
        (25,'8002','Almuerzos / Alimentación','otro_gasto'),
        (26,'8003','Compra Muebles y Enseres','otro_gasto'),
        (27,'8004','Otros','otro_gasto'),
        -- IMPUESTOS (códigos 9xxx — separados en ERI)
        (28,'9001','Impuestos','otro_gasto');
        CREATE TABLE IF NOT EXISTS losses (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            product_id INTEGER,
            product_name TEXT NOT NULL,
            quantity REAL NOT NULL,
            unit TEXT NOT NULL DEFAULT 'unidad',
            reason TEXT NOT NULL,
            responsible TEXT NOT NULL,
            notes TEXT NOT NULL DEFAULT '',
            created_at TEXT NOT NULL DEFAULT (datetime('now','localtime'))
        );
        CREATE TABLE IF NOT EXISTS cash_movements (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            type TEXT NOT NULL CHECK(type IN ('ingreso','egreso')),
            amount REAL NOT NULL,
            description TEXT NOT NULL,
            category TEXT NOT NULL DEFAULT 'General',
            created_at TEXT NOT NULL DEFAULT (datetime('now','localtime'))
        );
        CREATE TABLE IF NOT EXISTS workers (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            cargo TEXT NOT NULL DEFAULT '',
            phone TEXT NOT NULL DEFAULT '',
            active INTEGER NOT NULL DEFAULT 1,
            created_at TEXT NOT NULL DEFAULT (datetime('now','localtime'))
        );
        CREATE TABLE IF NOT EXISTS worker_notes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            worker_id INTEGER NOT NULL,
            content TEXT NOT NULL,
            created_at TEXT NOT NULL DEFAULT (datetime('now','localtime')),
            FOREIGN KEY (worker_id) REFERENCES workers(id)
        );
        CREATE TABLE IF NOT EXISTS tasks (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            funcion TEXT NOT NULL,
            area TEXT NOT NULL DEFAULT '',
            assigned_to TEXT NOT NULL DEFAULT '',
            status TEXT NOT NULL DEFAULT 'pendiente' CHECK(status IN ('pendiente','realizado')),
            created_at TEXT NOT NULL DEFAULT (datetime('now','localtime')),
            completed_at TEXT
        );
        CREATE TABLE IF NOT EXISTS checklist_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            turno TEXT NOT NULL CHECK(turno IN ('apertura','cierre')),
            section TEXT NOT NULL DEFAULT '',
            text TEXT NOT NULL,
            order_num INTEGER NOT NULL DEFAULT 0,
            active INTEGER NOT NULL DEFAULT 1
        );
        CREATE TABLE IF NOT EXISTS checklist_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            date TEXT NOT NULL,
            turno TEXT NOT NULL,
            item_id INTEGER NOT NULL,
            completed_by TEXT NOT NULL DEFAULT '',
            created_at TEXT NOT NULL DEFAULT (datetime('now','localtime')),
            UNIQUE(date, turno, item_id)
        );
        """)

        # Tabla de novedades
        conn.executescript("""
        CREATE TABLE IF NOT EXISTS novedades (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            tipo TEXT NOT NULL DEFAULT 'otro',
            descripcion TEXT NOT NULL,
            reportado_por TEXT NOT NULL DEFAULT '',
            created_at TEXT NOT NULL DEFAULT (datetime('now','localtime')),
            visto INTEGER NOT NULL DEFAULT 0
        );
        """)

        # Tabla de aseo semanal
        conn.executescript("""
        CREATE TABLE IF NOT EXISTS aseo_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            dia TEXT NOT NULL,
            text TEXT NOT NULL,
            order_num INTEGER NOT NULL DEFAULT 0,
            active INTEGER NOT NULL DEFAULT 1
        );
        CREATE TABLE IF NOT EXISTS aseo_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            date TEXT NOT NULL,
            dia TEXT NOT NULL,
            item_id INTEGER NOT NULL,
            completed_by TEXT NOT NULL DEFAULT '',
            created_at TEXT NOT NULL DEFAULT (datetime('now','localtime')),
            UNIQUE(date, dia, item_id)
        );
        """)

        # Tabla de documentos de trabajadores
        conn.execute("""
        CREATE TABLE IF NOT EXISTS worker_documents (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            worker_id INTEGER NOT NULL,
            filename TEXT NOT NULL DEFAULT '',
            original_name TEXT NOT NULL DEFAULT '',
            doc_type TEXT NOT NULL DEFAULT 'general',
            description TEXT NOT NULL DEFAULT '',
            created_at TEXT NOT NULL DEFAULT (datetime('now','localtime')),
            FOREIGN KEY (worker_id) REFERENCES workers(id)
        )""")

        # Tabla de cuentas abiertas (varias órdenes en paralelo)
        conn.execute("""
        CREATE TABLE IF NOT EXISTS open_accounts (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL DEFAULT 'Cuenta',
            items TEXT NOT NULL DEFAULT '[]',
            created_at TEXT NOT NULL DEFAULT (datetime('now','localtime')),
            updated_at TEXT NOT NULL DEFAULT (datetime('now','localtime'))
        )""")

        # Tabla de recepción de insumos
        conn.execute("""
        CREATE TABLE IF NOT EXISTS recepciones (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            descripcion TEXT NOT NULL,
            cantidad TEXT NOT NULL DEFAULT '',
            proveedor TEXT NOT NULL DEFAULT '',
            foto TEXT NOT NULL DEFAULT '',
            fecha_esperada TEXT NOT NULL DEFAULT '',
            estado TEXT NOT NULL DEFAULT 'pendiente',
            recibido_por TEXT NOT NULL DEFAULT '',
            recibido_at TEXT,
            nota_recepcion TEXT NOT NULL DEFAULT '',
            created_by TEXT NOT NULL DEFAULT '',
            created_at TEXT NOT NULL DEFAULT (datetime('now','localtime'))
        )""")

        # Agrega columna image si no existe (migración)
        try:
            conn.execute('ALTER TABLE products ADD COLUMN image TEXT DEFAULT NULL')
        except Exception:
            pass

        # Migración: agregar cost a products
        try:
            conn.execute('ALTER TABLE products ADD COLUMN cost REAL NOT NULL DEFAULT 0')
        except Exception:
            pass

        # Migración: agregar infinite_stock a products
        try:
            conn.execute('ALTER TABLE products ADD COLUMN infinite_stock INTEGER NOT NULL DEFAULT 0')
        except Exception:
            pass

        # Migración: agregar note_type a worker_notes
        try:
            conn.execute("ALTER TABLE worker_notes ADD COLUMN note_type TEXT NOT NULL DEFAULT 'nota'")
        except Exception:
            pass

        # Migración: agregar sale_value a losses
        try:
            conn.execute("ALTER TABLE losses ADD COLUMN sale_value REAL NOT NULL DEFAULT 0")
        except Exception:
            pass

        # Migración: agregar category a losses
        try:
            conn.execute("ALTER TABLE losses ADD COLUMN category TEXT NOT NULL DEFAULT ''")
        except Exception:
            pass

        # Seed aseo semanal (solo si la tabla está vacía)
        if conn.execute('SELECT COUNT(*) FROM aseo_items').fetchone()[0] == 0:
            aseo_seed = [
                ('lunes',     'Limpieza detrás de todos los electrónicos',                            1),
                ('lunes',     'Limpieza y organización de caja de almacenamiento',                    2),
                ('martes',    'Lavado de contenedores de azúcar y mezcladores',                       1),
                ('martes',    'Limpieza de equipos (sonido, cámaras, computador, caja de menuda)',    2),
                ('miercoles', 'Lavado de máquina de café',                                            1),
                ('jueves',    'Barrido y trapeado de pisos en áreas de cocina y atención al cliente', 1),
                ('viernes',   'Limpieza profunda del mesón de acero, cocina y máquinas (zona arriba y abajo)', 1),
                ('sabado',    'Limpieza de horno',                                                    1),
                ('sabado',    'Limpieza de nevera',                                                   2),
                ('sabado',    'Limpieza y organización de estantes de almacenamiento',                3),
            ]
            conn.executemany(
                'INSERT INTO aseo_items (dia, text, order_num) VALUES (?,?,?)',
                aseo_seed
            )

        # Seed checklist items (solo si la tabla está vacía)
        if conn.execute('SELECT COUNT(*) FROM checklist_items').fetchone()[0] == 0:
            apertura_items = [
                ('apertura', 'Antes de abrir', 'El turno anterior lo realizó correctamente', 1),
                ('apertura', 'Antes de abrir', 'Fritos y pasteles en exhibición', 2),
                ('apertura', 'Antes de abrir', 'Jugeras con insumo', 3),
                ('apertura', 'Antes de abrir', 'TVs y música prendidos', 4),
                ('apertura', 'Antes de abrir', 'Máquinas de café funcionando y con producto recargado', 5),
                ('apertura', 'Antes de abrir', 'Cortinas bien subidas', 6),
                ('apertura', 'Antes de abrir', 'Mesas organizadas y limpias', 7),
                ('apertura', 'Antes de abrir', 'Insumos recargados', 8),
                ('apertura', 'Antes de abrir', 'Caja base verificada y POS activo', 9),
                ('apertura', 'Antes de abrir', 'Vitrina y nevera bien surtidas', 10),
            ]
            cierre_items = [
                ('cierre', 'Limpieza y aseo', 'Lavar tanques de máquinas y jugeras', 1),
                ('cierre', 'Limpieza y aseo', 'Limpiar externamente la máquina de café y los hornos', 2),
                ('cierre', 'Limpieza y aseo', 'Lavar todos los elementos sucios — platos, charolas, tapetes de horneo, pinzas y utensilios', 3),
                ('cierre', 'Limpieza y aseo', 'Lavar y dejar desinfectados los trapos', 4),
                ('cierre', 'Limpieza y aseo', 'Limpiar la vitrina interna y externamente', 5),
                ('cierre', 'Limpieza y aseo', 'Barrer y trapear todo el local', 6),
                ('cierre', 'Limpieza y aseo', 'Entrar las mesas y sillas', 7),
                ('cierre', 'Producto e inventario', 'Guardar el producto sobrante en el congelador', 8),
                ('cierre', 'Producto e inventario', 'Registrar los sobrantes del día — producto, cantidad y estado', 9),
                ('cierre', 'Producto e inventario', 'Verificar que no quede nada fuera sin guardar', 10),
                ('cierre', 'Caja y sistema', 'Realizar el cierre de caja en el sistema', 11),
                ('cierre', 'Caja y sistema', 'Mandar la foto del cierre al administrador', 12),
                ('cierre', 'Caja y sistema', 'Verificar el efectivo y la caja de menuda', 13),
                ('cierre', 'Cierre del local', 'Apagar la máquina de café correctamente', 14),
                ('cierre', 'Cierre del local', 'Apagar los hornos y verificar que estén fríos', 15),
                ('cierre', 'Cierre del local', 'Apagar televisores, sonido y luces', 16),
                ('cierre', 'Cierre del local', 'Asegurar puertas, vitrinas y cajones con llave', 17),
                ('cierre', 'Registro final', 'Registrar novedades del turno — caja, inventario, equipos, clientes o personal', 18),
                ('cierre', 'Registro final', 'Llenar este checklist completamente', 19),
            ]
            conn.executemany(
                'INSERT INTO checklist_items (turno, section, text, order_num) VALUES (?,?,?,?)',
                apertura_items + cierre_items
            )

        # Seed trabajadores fijos
        for wname, wcargo in [('Camila', 'Empleada'), ('Daniela', 'Empleada'), ('Juan', 'Empleado')]:
            exists = conn.execute('SELECT id FROM workers WHERE name=?', (wname,)).fetchone()
            if not exists:
                conn.execute('INSERT INTO workers (name, cargo) VALUES (?,?)', (wname, wcargo))

        # Migración: reemplazar cuentas genéricas con cuentas específicas de la cafetería
        try:
            old_codes = conn.execute("SELECT code FROM accounts WHERE code IN ('4001','4002','5001','5002','6001','6002','6003','6004','6005','7001','7002','7003','7004','8001','9001','9002','9003')").fetchall()
            old_names = [r[0] for r in old_codes]
            # Verificar si aún tenemos las cuentas genéricas (por nombre de la primera)
            first_acc = conn.execute("SELECT name FROM accounts WHERE code='4001'").fetchone()
            if first_acc and first_acc[0] in ('Ventas de productos',):
                # Son cuentas genéricas — reemplazar si no hay journal_entries
                entry_count = conn.execute("SELECT COUNT(*) FROM journal_entries WHERE account_id <= 17").fetchone()[0]
                if entry_count == 0:
                    conn.execute("DELETE FROM accounts WHERE id <= 17")
            # Insertar/actualizar cuentas específicas
            new_accounts = [
                (1,'4001','Ventas Punto de Venta','ingreso'),
                (2,'4002','Ventas Institucional','ingreso'),
                (3,'5001','Compras Punto de Venta','costo'),
                (4,'5002','Cacaos','costo'),
                (5,'5003','Buñuelos','costo'),
                (6,'5004','Gaseosas y Desechables','costo'),
                (7,'5005','Fritos','costo'),
                (8,'5006','Otros Costos','costo'),
                (9,'6001','Nómina','gasto_operativo'),
                (10,'6002','Arriendo','gasto_operativo'),
                (11,'6003','Servicios Públicos','gasto_operativo'),
                (12,'6004','Internet','gasto_operativo'),
                (13,'6005','Prestaciones Sociales','gasto_operativo'),
                (14,'6006','Mercadeo','gasto_operativo'),
                (15,'6007','Fletes','gasto_operativo'),
                (16,'6008','Fletes Buñuelos','gasto_operativo'),
                (17,'6009','Fumigación','gasto_operativo'),
                (18,'6010','Dotación','gasto_operativo'),
                (19,'6011','Depreciación','gasto_operativo'),
                (20,'6012','Declaración','gasto_operativo'),
                (21,'6013','Licencia POS','gasto_operativo'),
                (22,'7001','Préstamo','gasto_admin'),
                (23,'7002','4x10000','gasto_admin'),
                (24,'8001','Compra de Equipos','otro_gasto'),
                (25,'8002','Almuerzos / Alimentación','otro_gasto'),
                (26,'8003','Compra Muebles y Enseres','otro_gasto'),
                (27,'8004','Otros','otro_gasto'),
                (28,'9001','Impuestos','otro_gasto'),
            ]
            for acc in new_accounts:
                conn.execute("INSERT OR IGNORE INTO accounts (id,code,name,type) VALUES (?,?,?,?)", acc)
                conn.execute("UPDATE accounts SET name=?, type=? WHERE id=? AND name != ?", (acc[2], acc[3], acc[0], acc[2]))
        except Exception as e:
            print('Account migration note:', e)
            pass

        count = conn.execute('SELECT COUNT(*) FROM products').fetchone()[0]
        if count == 0:
            samples = [
                ('Café americano','Bebidas',25,100,'taza',10),
                ('Café con leche','Bebidas',30,100,'taza',10),
                ('Capuchino','Bebidas',35,100,'taza',10),
                ('Té negro','Bebidas',20,50,'taza',10),
                ('Agua natural 500ml','Bebidas',15,24,'botella',6),
                ('Croissant','Panadería',28,20,'pieza',5),
                ('Muffin','Panadería',25,15,'pieza',5),
                ('Sándwich jamón','Comida',55,10,'pieza',3),
                ('Ensalada de frutas','Comida',45,8,'porción',3),
                ('Jugo naranja','Bebidas',35,10,'vaso',5),
            ]
            conn.executemany(
                'INSERT INTO products (name,category,price,stock,unit,low_stock_alert) VALUES (?,?,?,?,?,?)',
                samples
            )

def row_to_dict(row):
    return dict(row) if row else None

def rows_to_list(rows):
    return [dict(r) for r in rows]

# ─── STATIC ──────────────────────────────────────────
@app.route('/')
def index():
    return send_from_directory('public', 'index.html')

@app.route('/uploads/<path:path>')
def send_upload(path):
    return send_from_directory('public/uploads', path)

@app.route('/css/<path:path>')
def send_css(path):
    return send_from_directory('public/css', path)

@app.route('/js/<path:path>')
def send_js(path):
    return send_from_directory('public/js', path)

# ─── PRODUCTS ────────────────────────────────────────
@app.route('/api/products', methods=['GET'])
def get_products():
    with get_db() as conn:
        rows = conn.execute('SELECT * FROM products WHERE active=1 ORDER BY category, name').fetchall()
    return jsonify(rows_to_list(rows))

@app.route('/api/products', methods=['POST'])
def create_product():
    d = request.json
    if not d.get('name') or d.get('price') is None:
        return jsonify({'error': 'Nombre y precio requeridos'}), 400
    with get_db() as conn:
        cur = conn.execute(
            'INSERT INTO products (name,category,price,cost,stock,unit,low_stock_alert,infinite_stock) VALUES (?,?,?,?,?,?,?,?)',
            (d['name'], d.get('category','General'), d['price'], d.get('cost',0),
             d.get('stock',0), d.get('unit','unidad'), d.get('low_stock_alert',5), 1 if d.get('infinite_stock') else 0)
        )
        row = conn.execute('SELECT * FROM products WHERE id=?', (cur.lastrowid,)).fetchone()
    return jsonify(row_to_dict(row))

@app.route('/api/products/template', methods=['GET'])
def download_inventory_template():
    """Genera y descarga la plantilla Excel para importación masiva."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Inventario'

    # Estilos
    header_fill = PatternFill('solid', fgColor='FC4C02')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    center      = Alignment(horizontal='center', vertical='center')
    thin        = Side(border_style='thin', color='D1D5DB')
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ['Nombre *', 'Categoría', 'Precio Venta *', 'Precio Costo', 'Stock', 'Unidad', 'Alerta Stock Bajo', 'Stock Infinito (SI/NO)']
    widths  = [30, 16, 16, 14, 10, 12, 18, 20]

    for col, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = center
        cell.border    = border
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

    ws.row_dimensions[1].height = 28

    # Filas de ejemplo
    examples = [
        ('Café americano', 'Bebidas', 4500, 1200, 100, 'unidad', 10, 'NO'),
        ('Croissant mantequilla', 'Panadería', 3800, 900, 50, 'unidad', 5, 'NO'),
        ('Jugo de naranja', 'Bebidas', 5000, 1500, 0, 'unidad', 5, 'SI'),
        ('Agua natural 500ml', 'Bebidas', 2000, 600, 200, 'unidad', 20, 'NO'),
    ]
    eg_font = Font(size=10, color='374151')
    eg_fill = PatternFill('solid', fgColor='F9FAFB')
    eg_fill2 = PatternFill('solid', fgColor='FFFFFF')

    for row_i, ex in enumerate(examples, 2):
        fill = eg_fill if row_i % 2 == 0 else eg_fill2
        for col_i, val in enumerate(ex, 1):
            cell = ws.cell(row=row_i, column=col_i, value=val)
            cell.font   = eg_font
            cell.fill   = fill
            cell.border = border

    # Hoja de instrucciones
    ws2 = wb.create_sheet('Instrucciones')
    ws2['A1'] = '📋 INSTRUCCIONES DE IMPORTACIÓN'
    ws2['A1'].font = Font(bold=True, size=13, color='FC4C02')
    instrucciones = [
        ('', ''),
        ('Columna', 'Descripción'),
        ('Nombre *', 'Nombre del producto. OBLIGATORIO.'),
        ('Categoría', 'Categoría del producto (ej: Bebidas, Panadería). Default: General'),
        ('Precio Venta *', 'Precio de venta en pesos. OBLIGATORIO. Solo números.'),
        ('Precio Costo', 'Precio de costo/compra. Default: 0'),
        ('Stock', 'Cantidad en inventario. Default: 0'),
        ('Unidad', 'Unidad de medida (unidad, kg, litro, etc.). Default: unidad'),
        ('Alerta Stock Bajo', 'Cantidad mínima antes de alertar. Default: 5'),
        ('Stock Infinito (SI/NO)', 'SI = producto sin límite de stock. NO = controla stock. Default: NO'),
        ('', ''),
        ('⚠️ IMPORTANTE', 'No borrar ni mover la fila de encabezados (fila 1)'),
        ('⚠️ IMPORTANTE', 'Si el producto ya existe (mismo nombre exacto), se actualizará'),
        ('⚠️ IMPORTANTE', 'Si es nuevo, se creará automáticamente'),
    ]
    for r, (col_a, col_b) in enumerate(instrucciones, 2):
        ws2.cell(row=r, column=1, value=col_a).font = Font(bold=True, size=10)
        ws2.cell(row=r, column=2, value=col_b).font = Font(size=10)
    ws2.column_dimensions['A'].width = 28
    ws2.column_dimensions['B'].width = 60

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name='plantilla_inventario.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@app.route('/api/products/import', methods=['POST'])
def import_inventory():
    """Importa productos desde un archivo Excel."""
    if 'file' not in request.files:
        return jsonify({'error': 'No se recibió archivo'}), 400
    f = request.files['file']
    if not f.filename.lower().endswith(('.xlsx', '.xls')):
        return jsonify({'error': 'Solo se aceptan archivos .xlsx o .xls'}), 400

    try:
        wb = openpyxl.load_workbook(io.BytesIO(f.read()), data_only=True)
        ws = wb.active
    except Exception as e:
        return jsonify({'error': f'No se pudo leer el archivo: {str(e)}'}), 400

    # Detectar columnas por encabezado (fila 1)
    headers = {}
    for col in ws.iter_cols(min_row=1, max_row=1):
        cell = col[0]
        if cell.value:
            key = str(cell.value).strip().lower()
            # Normalizar
            if 'nombre' in key:           headers['name']             = cell.column - 1
            elif 'categorí' in key or 'categoria' in key: headers['category'] = cell.column - 1
            elif 'venta' in key or 'precio v' in key:     headers['price']    = cell.column - 1
            elif 'costo' in key or 'precio c' in key:     headers['cost']     = cell.column - 1
            elif 'stock bajo' in key or 'alerta' in key:  headers['low_stock_alert'] = cell.column - 1
            elif 'infinito' in key:        headers['infinite_stock']   = cell.column - 1
            elif 'stock' in key:           headers['stock']            = cell.column - 1
            elif 'unidad' in key:          headers['unit']             = cell.column - 1

    if 'name' not in headers or 'price' not in headers:
        return jsonify({'error': 'El archivo debe tener columnas "Nombre" y "Precio Venta"'}), 400

    created = 0
    updated = 0
    errors  = []

    with get_db() as conn:
        for row_i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            if not any(row):
                continue  # fila vacía
            def col(k, default=None):
                idx = headers.get(k)
                return row[idx] if idx is not None and idx < len(row) else default

            name = str(col('name', '')).strip() if col('name') else ''
            if not name:
                continue

            try:
                price    = float(str(col('price', 0)).replace(',','.').replace(' ','') or 0)
            except:
                errors.append(f'Fila {row_i}: precio inválido para "{name}"')
                continue

            category = str(col('category', 'General') or 'General').strip()
            try:
                cost = float(str(col('cost', 0) or 0).replace(',','.').replace(' ',''))
            except: cost = 0
            try:
                stock = int(float(str(col('stock', 0) or 0).replace(',','.').replace(' ','')))
            except: stock = 0
            unit = str(col('unit', 'unidad') or 'unidad').strip()
            try:
                low_stock_alert = int(float(str(col('low_stock_alert', 5) or 5)))
            except: low_stock_alert = 5
            inf_val = str(col('infinite_stock', 'NO') or 'NO').strip().upper()
            infinite_stock = 1 if inf_val in ('SI', 'SÍ', 'YES', '1', 'TRUE') else 0

            existing = conn.execute(
                'SELECT id FROM products WHERE LOWER(name)=LOWER(?) AND active=1', (name,)
            ).fetchone()

            if existing:
                conn.execute(
                    'UPDATE products SET category=?,price=?,cost=?,stock=?,unit=?,low_stock_alert=?,infinite_stock=? WHERE id=?',
                    (category, price, cost, stock, unit, low_stock_alert, infinite_stock, existing['id'])
                )
                updated += 1
            else:
                conn.execute(
                    'INSERT INTO products (name,category,price,cost,stock,unit,low_stock_alert,infinite_stock) VALUES (?,?,?,?,?,?,?,?)',
                    (name, category, price, cost, stock, unit, low_stock_alert, infinite_stock)
                )
                created += 1

    return jsonify({
        'ok': True,
        'created': created,
        'updated': updated,
        'errors': errors,
        'total': created + updated
    })

@app.route('/api/products/<int:pid>/image', methods=['POST'])
def upload_product_image(pid):
    if 'image' not in request.files:
        return jsonify({'error': 'No se recibió imagen'}), 400
    file = request.files['image']
    if not file.filename:
        return jsonify({'error': 'No se seleccionó archivo'}), 400
    ext = file.filename.rsplit('.', 1)[-1].lower()
    if ext not in ('jpg', 'jpeg', 'png', 'gif', 'webp'):
        return jsonify({'error': 'Formato no permitido. Usa jpg, png, gif o webp'}), 400
    upload_dir = os.path.join(os.path.dirname(__file__), 'public', 'uploads', 'products')
    os.makedirs(upload_dir, exist_ok=True)
    filename = f'product_{pid}.{ext}'
    # Eliminar imagen anterior con distinta extensión
    for old_ext in ('jpg', 'jpeg', 'png', 'gif', 'webp'):
        old_path = os.path.join(upload_dir, f'product_{pid}.{old_ext}')
        if os.path.exists(old_path) and old_ext != ext:
            os.remove(old_path)
    file.save(os.path.join(upload_dir, filename))
    with get_db() as conn:
        conn.execute('UPDATE products SET image=? WHERE id=?', (filename, pid))
        row = conn.execute('SELECT * FROM products WHERE id=?', (pid,)).fetchone()
    return jsonify(row_to_dict(row))

@app.route('/api/products/<int:pid>', methods=['PUT'])
def update_product(pid):
    d = request.json
    with get_db() as conn:
        conn.execute(
            'UPDATE products SET name=?,category=?,price=?,cost=?,stock=?,unit=?,low_stock_alert=?,infinite_stock=? WHERE id=?',
            (d['name'],d['category'],d['price'],d.get('cost',0),d['stock'],d['unit'],d['low_stock_alert'],1 if d.get('infinite_stock') else 0,pid)
        )
        row = conn.execute('SELECT * FROM products WHERE id=?', (pid,)).fetchone()
    return jsonify(row_to_dict(row))

@app.route('/api/products/<int:pid>', methods=['DELETE'])
def delete_product(pid):
    with get_db() as conn:
        conn.execute('UPDATE products SET active=0 WHERE id=?', (pid,))
    return jsonify({'ok': True})

# ─── SALES ───────────────────────────────────────────
@app.route('/api/sales', methods=['GET'])
def get_sales():
    from_d = request.args.get('from')
    to_d = request.args.get('to')
    with get_db() as conn:
        if from_d and to_d:
            rows = conn.execute(
                "SELECT * FROM sales WHERE date(created_at) BETWEEN date(?) AND date(?) ORDER BY created_at DESC",
                (from_d, to_d)
            ).fetchall()
        else:
            rows = conn.execute('SELECT * FROM sales ORDER BY created_at DESC').fetchall()
    return jsonify(rows_to_list(rows))

@app.route('/api/sales/<int:sid>', methods=['GET'])
def get_sale(sid):
    with get_db() as conn:
        sale = conn.execute('SELECT * FROM sales WHERE id=?', (sid,)).fetchone()
        if not sale:
            return jsonify({'error': 'Venta no encontrada'}), 404
        items = conn.execute('SELECT * FROM sale_items WHERE sale_id=?', (sid,)).fetchall()
    result = row_to_dict(sale)
    result['items'] = rows_to_list(items)
    return jsonify(result)

@app.route('/api/sales', methods=['POST'])
def create_sale():
    d = request.json
    items = d.get('items', [])
    if not items:
        return jsonify({'error': 'No hay productos en la venta'}), 400

    subtotal = sum(i['price'] * i['quantity'] for i in items)
    total = subtotal

    with get_db() as conn:
        cur = conn.execute(
            'INSERT INTO sales (subtotal,total,payment_method,notes) VALUES (?,?,?,?)',
            (subtotal, total, d.get('payment_method','efectivo'), d.get('notes',''))
        )
        sale_id = cur.lastrowid

        for item in items:
            conn.execute(
                'INSERT INTO sale_items (sale_id,product_id,product_name,price,quantity,subtotal) VALUES (?,?,?,?,?,?)',
                (sale_id, item.get('product_id'), item['product_name'],
                 item['price'], item['quantity'], item['price'] * item['quantity'])
            )
            if item.get('product_id'):
                conn.execute(
                    'UPDATE products SET stock = MAX(0, stock - ?) WHERE id=? AND infinite_stock=0',
                    (item['quantity'], item['product_id'])
                )

        conn.execute(
            "INSERT INTO cash_movements (type,amount,description,category) VALUES ('ingreso',?,?,?)",
            (total, f'Venta #{sale_id}', 'Ventas')
        )

        sale = conn.execute('SELECT * FROM sales WHERE id=?', (sale_id,)).fetchone()
        sale_items = conn.execute('SELECT * FROM sale_items WHERE sale_id=?', (sale_id,)).fetchall()

    result = row_to_dict(sale)
    result['items'] = rows_to_list(sale_items)
    return jsonify(result)

# ─── MOVEMENTS ───────────────────────────────────────
@app.route('/api/movements', methods=['GET'])
def get_movements():
    from_d = request.args.get('from')
    to_d = request.args.get('to')
    with get_db() as conn:
        if from_d and to_d:
            rows = conn.execute(
                "SELECT * FROM cash_movements WHERE date(created_at) BETWEEN date(?) AND date(?) ORDER BY created_at DESC",
                (from_d, to_d)
            ).fetchall()
        else:
            rows = conn.execute('SELECT * FROM cash_movements ORDER BY created_at DESC').fetchall()
    return jsonify(rows_to_list(rows))

@app.route('/api/movements', methods=['POST'])
def create_movement():
    d = request.json
    if not d.get('type') or not d.get('amount') or not d.get('description'):
        return jsonify({'error': 'Tipo, monto y descripción requeridos'}), 400
    with get_db() as conn:
        cur = conn.execute(
            'INSERT INTO cash_movements (type,amount,description,category) VALUES (?,?,?,?)',
            (d['type'], d['amount'], d['description'], d.get('category','General'))
        )
        row = conn.execute('SELECT * FROM cash_movements WHERE id=?', (cur.lastrowid,)).fetchone()
    return jsonify(row_to_dict(row))

@app.route('/api/movements/<int:mid>', methods=['DELETE'])
def delete_movement(mid):
    with get_db() as conn:
        conn.execute('DELETE FROM cash_movements WHERE id=?', (mid,))
    return jsonify({'ok': True})

# ─── DASHBOARD ───────────────────────────────────────
@app.route('/api/dashboard', methods=['GET'])
def dashboard():
    today = datetime.now().strftime('%Y-%m-%d')
    with get_db() as conn:
        ventas = conn.execute(
            "SELECT COUNT(*) as count, COALESCE(SUM(total),0) as total FROM sales WHERE date(created_at)=date(?)",
            (today,)
        ).fetchone()
        ingreso = conn.execute(
            "SELECT COALESCE(SUM(amount),0) as total FROM cash_movements WHERE type='ingreso' AND date(created_at)=date(?)",
            (today,)
        ).fetchone()
        egreso = conn.execute(
            "SELECT COALESCE(SUM(amount),0) as total FROM cash_movements WHERE type='egreso' AND date(created_at)=date(?)",
            (today,)
        ).fetchone()
        bajo_stock = conn.execute(
            'SELECT * FROM products WHERE active=1 AND stock<=low_stock_alert ORDER BY stock ASC LIMIT 10'
        ).fetchall()
        top_productos = conn.execute("""
            SELECT si.product_name, SUM(si.quantity) as qty, SUM(si.subtotal) as total
            FROM sale_items si JOIN sales s ON s.id=si.sale_id
            WHERE date(s.created_at) >= date('now','-7 days')
            GROUP BY si.product_name ORDER BY total DESC LIMIT 5
        """).fetchall()
        ventas_semana = conn.execute("""
            SELECT date(created_at) as dia, COUNT(*) as ventas, SUM(total) as total
            FROM sales WHERE date(created_at) >= date('now','-6 days')
            GROUP BY dia ORDER BY dia
        """).fetchall()

        # Ventas por hora del día de hoy (0-23)
        ventas_hora_raw = conn.execute("""
            SELECT CAST(strftime('%H', created_at) AS INTEGER) as hora,
                   COUNT(*) as ventas,
                   COALESCE(SUM(total), 0) as total
            FROM sales
            WHERE date(created_at) = date(?)
            GROUP BY hora ORDER BY hora
        """, (today,)).fetchall()

    # Rellenar las 24 horas (0-23) con 0 si no hay datos
    hora_map = {r['hora']: {'ventas': r['ventas'], 'total': r['total']} for r in ventas_hora_raw}
    ventas_por_hora = [
        {'hora': h, 'ventas': hora_map.get(h, {}).get('ventas', 0), 'total': hora_map.get(h, {}).get('total', 0)}
        for h in range(24)
    ]

    return jsonify({
        'ventasHoy': ventas['count'],
        'totalHoy': ventas['total'],
        'ingresoHoy': ingreso['total'],
        'egresoHoy': egreso['total'],
        'balanceHoy': ingreso['total'] - egreso['total'],
        'bajoStock': rows_to_list(bajo_stock),
        'topProductos': rows_to_list(top_productos),
        'ventasSemana': rows_to_list(ventas_semana),
        'ventasPorHora': ventas_por_hora,
    })

# ─── REPORTES ────────────────────────────────────────────────────────────────
def _wb_style():
    """Retorna funciones de estilo reutilizables."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    ORANGE  = 'FC4C02'; DARK = '1E293B'; LIGHT = 'F8FAFC'; GRAY = 'E2E8F0'
    PURPLE  = '6366F1'; GREEN = '16A34A'; RED = 'DC2626'
    thin = Side(border_style='thin', color='D1D5DB')
    brd  = Border(left=thin, right=thin, top=thin, bottom=thin)
    def hdr(ws, row, cols, fill_hex=ORANGE, font_hex='FFFFFF', height=26):
        fill = PatternFill('solid', fgColor=fill_hex)
        font = Font(bold=True, color=font_hex, size=11)
        align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        for col, val in enumerate(cols, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.fill = fill; c.font = font; c.alignment = align; c.border = brd
        ws.row_dimensions[row].height = height
    def cell(ws, r, c, val, bold=False, color=DARK, align='left', fmt=None, fill_hex=None):
        cl = ws.cell(row=r, column=c, value=val)
        cl.font = Font(size=10, bold=bold, color=color)
        cl.alignment = Alignment(horizontal=align, vertical='center')
        cl.border = brd
        if fill_hex: cl.fill = PatternFill('solid', fgColor=fill_hex)
        if fmt:      cl.number_format = fmt
        return cl
    def title(ws, text, subtitle=''):
        ws['A1'] = text; ws['A1'].font = Font(bold=True, size=16, color=ORANGE)
        ws['A1'].alignment = Alignment(horizontal='left')
        if subtitle:
            ws['A2'] = subtitle; ws['A2'].font = Font(size=11, color='64748B')
        return 3 if subtitle else 2
    def col_w(ws, widths):
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    def fmt_cop(v): return f"${v:,.0f}".replace(',','.')
    return hdr, cell, title, col_w, fmt_cop, ORANGE, DARK, GRAY, PURPLE, GREEN, RED, brd

@app.route('/api/reports/daily', methods=['GET'])
def report_daily():
    date = request.args.get('date', datetime.now().strftime('%Y-%m-%d'))
    hdr, cell, title_fn, col_w, fmt_cop, ORANGE, DARK, GRAY, PURPLE, GREEN, RED, brd = _wb_style()
    wb = openpyxl.Workbook()

    with get_db() as conn:
        # ── RESUMEN ──
        ventas = conn.execute("SELECT COUNT(*) c, COALESCE(SUM(total),0) t, COALESCE(SUM(subtotal),0) s FROM sales WHERE date(created_at)=?", (date,)).fetchone()
        ingresos = conn.execute("SELECT COALESCE(SUM(amount),0) t FROM cash_movements WHERE type='ingreso' AND date(created_at)=?", (date,)).fetchone()
        egresos  = conn.execute("SELECT COALESCE(SUM(amount),0) t FROM cash_movements WHERE type='egreso'  AND date(created_at)=?", (date,)).fetchone()
        cierre   = conn.execute("SELECT * FROM cash_registers WHERE date(opened_at)=? ORDER BY opened_at DESC LIMIT 1", (date,)).fetchone()

        # ── VENTAS POR PRODUCTO ──
        productos = conn.execute("""SELECT si.product_name, SUM(si.quantity) qty, SUM(si.subtotal) subtotal
            FROM sale_items si JOIN sales s ON s.id=si.sale_id
            WHERE date(s.created_at)=? GROUP BY si.product_name ORDER BY subtotal DESC""", (date,)).fetchall()

        # ── VENTAS POR HORA ──
        por_hora = conn.execute("""SELECT CAST(strftime('%H', created_at) AS INTEGER) h,
            COUNT(*) c, COALESCE(SUM(total),0) t FROM sales WHERE date(created_at)=?
            GROUP BY h ORDER BY h""", (date,)).fetchall()

        # ── EGRESOS ──
        movs_egreso = conn.execute("""SELECT description, amount, created_at FROM cash_movements
            WHERE type='egreso' AND date(created_at)=? ORDER BY created_at""", (date,)).fetchall()

        # ── TAREAS ──
        tareas = conn.execute("""SELECT funcion, area, assigned_to, status, created_at FROM tasks
            WHERE date(created_at)=? ORDER BY status""", (date,)).fetchall()

        # ── PÉRDIDAS ──
        perdidas = conn.execute("""SELECT product_name, category, quantity, sale_value, responsible, reason, created_at
            FROM losses WHERE date(created_at)=? ORDER BY created_at""", (date,)).fetchall()

        # ── CHECKLIST ──
        cl_apertura = conn.execute("""SELECT ci.text, cl.completed_by, cl.created_at FROM checklist_items ci
            LEFT JOIN checklist_logs cl ON cl.item_id=ci.id AND cl.date=? AND cl.turno='apertura'
            WHERE ci.turno='apertura' AND ci.active=1 ORDER BY ci.order_num""", (date,)).fetchall()
        cl_cierre = conn.execute("""SELECT ci.text, cl.completed_by, cl.created_at FROM checklist_items ci
            LEFT JOIN checklist_logs cl ON cl.item_id=ci.id AND cl.date=? AND cl.turno='cierre'
            WHERE ci.turno='cierre' AND ci.active=1 ORDER BY ci.order_num""", (date,)).fetchall()

        # ── NOVEDADES ──
        novedades = conn.execute("""SELECT tipo, descripcion, reportado_por, created_at FROM novedades
            WHERE date(created_at)=? ORDER BY created_at""", (date,)).fetchall()

    # Hoja 1 — Resumen
    ws = wb.active; ws.title = '📊 Resumen'
    col_w(ws, [28, 18, 18, 18, 18])
    start = title_fn(ws, f'Informe Diario — {date}', f'Generado el {datetime.now().strftime("%Y-%m-%d %H:%M")}')
    r = start
    hdr(ws, r, ['INDICADOR', 'VALOR']); r += 1
    datos_resumen = [
        ('Total tickets vendidos', ventas['c']),
        ('Total ventas (bruto)',   ventas['t']),
        ('Total ingresos en caja', ingresos['t']),
        ('Total egresos en caja',  egresos['t']),
        ('Balance del día',        ingresos['t'] - egresos['t']),
    ]
    if cierre:
        datos_resumen += [
            ('Base apertura caja', cierre['opening_balance'] if cierre else 0),
            ('Diferencia cierre',  cierre['difference'] if cierre else 0),
        ]
    fills = ['FFFFFF','F8FAFC']
    for i, (k, v) in enumerate(datos_resumen):
        fill = fills[i % 2]
        cell(ws, r, 1, k, bold=True, fill_hex=fill)
        is_money = isinstance(v, float) or (isinstance(v, int) and 'Total' in k or 'Base' in k or 'Difer' in k or 'ingreso' in k or 'egreso' in k or 'Balance' in k)
        cell(ws, r, 2, v, bold=True, color=ORANGE if 'Balance' in k else DARK, align='right', fill_hex=fill)
        ws.cell(r, 2).number_format = '$#,##0'
        r += 1

    # Hoja 2 — Productos
    ws2 = wb.create_sheet('🛒 Productos')
    col_w(ws2, [32, 12, 18, 14])
    title_fn(ws2, 'Ventas por producto', date)
    r2 = 3
    hdr(ws2, r2, ['PRODUCTO', 'CANTIDAD', 'SUBTOTAL', '% DEL TOTAL']); r2 += 1
    total_vtas = sum(p['subtotal'] for p in productos) or 1
    for i, p in enumerate(productos):
        fill = 'FFFFFF' if i%2==0 else 'F8FAFC'
        pct = round(p['subtotal'] / total_vtas * 100, 1)
        cell(ws2, r2, 1, p['product_name'], fill_hex=fill)
        cell(ws2, r2, 2, int(p['qty']), align='center', fill_hex=fill)
        cell(ws2, r2, 3, p['subtotal'], align='right', fill_hex=fill); ws2.cell(r2,3).number_format='$#,##0'
        cell(ws2, r2, 4, f'{pct}%', align='center', fill_hex=fill)
        r2 += 1
    if productos:
        hdr(ws2, r2, ['TOTAL', int(sum(p['qty'] for p in productos)), ventas['s'], '100%'], fill_hex='1E293B')
        ws2.cell(r2,3).number_format='$#,##0'

    # Hoja 3 — Por hora
    ws3 = wb.create_sheet('⏰ Por hora')
    col_w(ws3, [14, 12, 18])
    title_fn(ws3, 'Comportamiento de ventas por hora', date)
    r3 = 3
    hdr(ws3, r3, ['HORA', 'TICKETS', 'INGRESOS']); r3 += 1
    hora_map = {p['h']: p for p in por_hora}
    for h in range(7, 22):
        p = hora_map.get(h, {'c': 0, 't': 0})
        fill = 'FFF3EE' if p['c'] == max((x['c'] for x in por_hora), default=0) and p['c'] > 0 else ('FFFFFF' if h%2==0 else 'F8FAFC')
        cell(ws3, r3, 1, f'{h:02d}:00', align='center', fill_hex=fill)
        cell(ws3, r3, 2, p['c'], align='center', fill_hex=fill)
        cell(ws3, r3, 3, p['t'], align='right', fill_hex=fill); ws3.cell(r3,3).number_format='$#,##0'
        r3 += 1

    # Hoja 4 — Egresos
    ws4 = wb.create_sheet('💸 Egresos')
    col_w(ws4, [36, 16, 18])
    title_fn(ws4, 'Egresos del día', date)
    r4 = 3
    hdr(ws4, r4, ['CONCEPTO', 'MONTO', 'HORA']); r4 += 1
    for i, e in enumerate(movs_egreso):
        fill = 'FFFFFF' if i%2==0 else 'F8FAFC'
        cell(ws4, r4, 1, e['description'] or '—', fill_hex=fill)
        cell(ws4, r4, 2, e['amount'], align='right', fill_hex=fill, color=RED); ws4.cell(r4,2).number_format='$#,##0'
        cell(ws4, r4, 3, str(e['created_at'])[11:16], align='center', fill_hex=fill)
        r4 += 1
    if movs_egreso:
        hdr(ws4, r4, ['TOTAL', egresos['t'], ''], fill_hex='1E293B')
        ws4.cell(r4,2).number_format='$#,##0'

    # Hoja 5 — Checklist
    ws5 = wb.create_sheet('✅ Checklist')
    col_w(ws5, [42, 12, 18])
    title_fn(ws5, 'Checklist del día', date)
    r5 = 3
    for turno_label, items in [('🌅 APERTURA', cl_apertura), ('🌆 CIERRE', cl_cierre)]:
        hdr(ws5, r5, [turno_label, 'ESTADO', 'COMPLETADO POR'], fill_hex='334155'); r5 += 1
        completados = sum(1 for x in items if x['completed_by'])
        for i, it in enumerate(items):
            done = bool(it['completed_by'])
            fill = 'F0FDF4' if done else 'FFFFFF' if i%2==0 else 'F8FAFC'
            cell(ws5, r5, 1, it['text'], fill_hex=fill)
            cell(ws5, r5, 2, '✅ Hecho' if done else '⬜ Pendiente', align='center', fill_hex=fill, color=GREEN if done else '94A3B8')
            cell(ws5, r5, 3, it['completed_by'] or '—', align='center', fill_hex=fill)
            r5 += 1
        hdr(ws5, r5, [f'Completados: {completados}/{len(items)}', f'{round(completados/len(items)*100) if items else 0}%', ''], fill_hex='E2E8F0', font_hex='374151'); r5 += 2

    # Hoja 6 — Novedades & Pérdidas
    ws6 = wb.create_sheet('📢 Novedades')
    col_w(ws6, [16, 42, 16, 16])
    title_fn(ws6, 'Novedades y Pérdidas', date)
    r6 = 3
    hdr(ws6, r6, ['TIPO', 'DESCRIPCIÓN', 'REPORTADO POR', 'HORA']); r6 += 1
    for i, n in enumerate(novedades):
        fill = 'FFFFFF' if i%2==0 else 'F8FAFC'
        cell(ws6, r6, 1, n['tipo'].upper(), align='center', fill_hex=fill)
        cell(ws6, r6, 2, n['descripcion'], fill_hex=fill)
        cell(ws6, r6, 3, n['reportado_por'], align='center', fill_hex=fill)
        cell(ws6, r6, 4, str(n['created_at'])[11:16], align='center', fill_hex=fill)
        r6 += 1
    if not novedades:
        cell(ws6, r6, 1, 'Sin novedades registradas', color='94A3B8')
    r6 += 1
    hdr(ws6, r6, ['PRODUCTO', 'CATEGORÍA', 'CANTIDAD', 'VALOR VENTA', 'RESPONSABLE', 'MOTIVO'], fill_hex='334155'); r6 += 1
    ws6.column_dimensions['E'].width = 16; ws6.column_dimensions['F'].width = 28
    for i, p in enumerate(perdidas):
        fill = 'FEF2F2' if i%2==0 else 'FFFFFF'
        cell(ws6, r6, 1, p['product_name'], fill_hex=fill)
        cell(ws6, r6, 2, p['category'] or '—', align='center', fill_hex=fill)
        cell(ws6, r6, 3, p['quantity'], align='center', fill_hex=fill)
        cell(ws6, r6, 4, p['sale_value'] or 0, align='right', fill_hex=fill, color=RED); ws6.cell(r6,4).number_format='$#,##0'
        cell(ws6, r6, 5, p['responsible'] or '—', align='center', fill_hex=fill)
        cell(ws6, r6, 6, p['reason'] or '—', fill_hex=fill)
        r6 += 1
    if not perdidas: cell(ws6, r6, 1, 'Sin pérdidas registradas', color='94A3B8')

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    fname = f'informe_diario_{date}.xlsx'
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/api/reports/weekly', methods=['GET'])
def report_weekly():
    date_str = request.args.get('date', datetime.now().strftime('%Y-%m-%d'))
    hdr, cell, title_fn, col_w, fmt_cop, ORANGE, DARK, GRAY, PURPLE, GREEN, RED, brd = _wb_style()
    wb = openpyxl.Workbook()

    with get_db() as conn:
        # Semana: lunes a domingo del periodo
        from datetime import timedelta
        ref = datetime.strptime(date_str, '%Y-%m-%d')
        lunes = ref - timedelta(days=ref.weekday())
        domingo = lunes + timedelta(days=6)
        d_ini = lunes.strftime('%Y-%m-%d'); d_fin = domingo.strftime('%Y-%m-%d')

        dias_semana = conn.execute("""SELECT date(created_at) dia, COUNT(*) tickets, COALESCE(SUM(total),0) total
            FROM sales WHERE date(created_at) BETWEEN ? AND ? GROUP BY dia ORDER BY dia""", (d_ini, d_fin)).fetchall()
        egresos_sem = conn.execute("""SELECT date(created_at) dia, COALESCE(SUM(amount),0) total
            FROM cash_movements WHERE type='egreso' AND date(created_at) BETWEEN ? AND ? GROUP BY dia ORDER BY dia""", (d_ini, d_fin)).fetchall()
        ingresos_sem = conn.execute("""SELECT date(created_at) dia, COALESCE(SUM(amount),0) total
            FROM cash_movements WHERE type='ingreso' AND date(created_at) BETWEEN ? AND ? GROUP BY dia ORDER BY dia""", (d_ini, d_fin)).fetchall()
        top_prods = conn.execute("""SELECT si.product_name, SUM(si.quantity) qty, SUM(si.subtotal) total
            FROM sale_items si JOIN sales s ON s.id=si.sale_id
            WHERE date(s.created_at) BETWEEN ? AND ? GROUP BY si.product_name ORDER BY total DESC LIMIT 20""", (d_ini, d_fin)).fetchall()
        por_hora_sem = conn.execute("""SELECT CAST(strftime('%H', created_at) AS INTEGER) h,
            COUNT(*) tickets, COALESCE(SUM(total),0) total
            FROM sales WHERE date(created_at) BETWEEN ? AND ?
            GROUP BY h ORDER BY h""", (d_ini, d_fin)).fetchall()
        cierres = conn.execute("""SELECT opened_at, closed_at, opening_balance, closing_balance, difference, notes
            FROM cash_registers WHERE date(opened_at) BETWEEN ? AND ? ORDER BY opened_at""", (d_ini, d_fin)).fetchall()
        perdidas_sem = conn.execute("""SELECT product_name, category, quantity, sale_value, responsible, reason, created_at
            FROM losses WHERE date(created_at) BETWEEN ? AND ? ORDER BY created_at""", (d_ini, d_fin)).fetchall()
        novedades_sem = conn.execute("""SELECT tipo, descripcion, reportado_por, created_at
            FROM novedades WHERE date(created_at) BETWEEN ? AND ? ORDER BY created_at""", (d_ini, d_fin)).fetchall()
        tareas_sem = conn.execute("""SELECT funcion, area, assigned_to, status FROM tasks
            WHERE date(created_at) BETWEEN ? AND ?""", (d_ini, d_fin)).fetchall()
        workers_perdidas = conn.execute("""SELECT responsible, COUNT(*) cnt, COALESCE(SUM(sale_value),0) total
            FROM losses WHERE date(created_at) BETWEEN ? AND ? AND responsible IS NOT NULL AND responsible!=''
            GROUP BY responsible ORDER BY total DESC""", (d_ini, d_fin)).fetchall()

    semana_label = f'{d_ini} → {d_fin}'

    # Hoja 1 — Resumen semanal
    ws = wb.active; ws.title = '📊 Resumen'
    col_w(ws, [16, 12, 18, 16, 16, 16, 16])
    title_fn(ws, f'Informe Semanal — {semana_label}', f'Generado el {datetime.now().strftime("%Y-%m-%d %H:%M")}')
    r = 3
    hdr(ws, r, ['DÍA', 'TICKETS', 'VENTAS', 'INGRESOS', 'EGRESOS', 'BALANCE', 'VARIACIÓN']); r += 1
    egr_map = {e['dia']: e['total'] for e in egresos_sem}
    ing_map = {e['dia']: e['total'] for e in ingresos_sem}
    dias_nombres = {'Monday':'Lunes','Tuesday':'Martes','Wednesday':'Miércoles','Thursday':'Jueves','Friday':'Viernes','Saturday':'Sábado','Sunday':'Domingo'}
    total_tickets = total_ventas = total_ingresos = total_egresos = 0
    prev_total = None
    for i in range(7):
        d = (lunes + timedelta(days=i)).strftime('%Y-%m-%d')
        dia_nom = dias_nombres.get((lunes + timedelta(days=i)).strftime('%A'), d)
        dia_data = next((x for x in dias_semana if x['dia'] == d), None)
        t = dia_data['tickets'] if dia_data else 0
        v = dia_data['total']   if dia_data else 0
        ing = ing_map.get(d, 0); egr = egr_map.get(d, 0); bal = ing - egr
        var = ''
        if prev_total is not None and prev_total > 0: var = f'{round((v-prev_total)/prev_total*100,1)}%'
        fill = 'FFF3EE' if v == max((x['total'] for x in dias_semana), default=0) and v > 0 else ('FFFFFF' if i%2==0 else 'F8FAFC')
        cell(ws, r, 1, f'{dia_nom} {d[8:]}', bold=True, fill_hex=fill)
        cell(ws, r, 2, t, align='center', fill_hex=fill)
        cell(ws, r, 3, v, align='right', fill_hex=fill); ws.cell(r,3).number_format='$#,##0'
        cell(ws, r, 4, ing, align='right', fill_hex=fill); ws.cell(r,4).number_format='$#,##0'
        cell(ws, r, 5, egr, align='right', fill_hex=fill, color=RED); ws.cell(r,5).number_format='$#,##0'
        cell(ws, r, 6, bal, align='right', fill_hex=fill, color=GREEN if bal>=0 else RED, bold=True); ws.cell(r,6).number_format='$#,##0'
        cell(ws, r, 7, var, align='center', fill_hex=fill)
        total_tickets+=t; total_ventas+=v; total_ingresos+=ing; total_egresos+=egr; prev_total=v; r+=1
    hdr(ws, r, ['TOTAL SEMANA', total_tickets, total_ventas, total_ingresos, total_egresos, total_ingresos-total_egresos, ''], fill_hex='1E293B')
    for c in [3,4,5,6]: ws.cell(r,c).number_format='$#,##0'

    # Hoja 2 — Top productos
    ws2 = wb.create_sheet('🛒 Top Productos')
    col_w(ws2, [34, 12, 18, 14, 12])
    title_fn(ws2, 'Productos más vendidos', semana_label)
    r2 = 3
    hdr(ws2, r2, ['PRODUCTO', 'CANTIDAD', 'TOTAL', '% PARTICIPACIÓN', 'RANKING']); r2 += 1
    tot_p = sum(p['total'] for p in top_prods) or 1
    for i, p in enumerate(top_prods):
        fill = 'FFF3EE' if i==0 else 'FFFFFF' if i%2==0 else 'F8FAFC'
        pct = round(p['total']/tot_p*100,1)
        cell(ws2, r2, 1, p['product_name'], bold=(i==0), fill_hex=fill)
        cell(ws2, r2, 2, int(p['qty']), align='center', fill_hex=fill)
        cell(ws2, r2, 3, p['total'], align='right', fill_hex=fill); ws2.cell(r2,3).number_format='$#,##0'
        cell(ws2, r2, 4, f'{pct}%', align='center', fill_hex=fill)
        medal = ['🥇','🥈','🥉'] ; cell(ws2, r2, 5, medal[i] if i<3 else str(i+1), align='center', fill_hex=fill)
        r2 += 1

    # Hoja 3 — Comportamiento por hora
    ws3 = wb.create_sheet('⏰ Horas Pico')
    col_w(ws3, [14, 14, 18])
    title_fn(ws3, 'Comportamiento por hora (semana)', semana_label)
    r3 = 3; hdr(ws3, r3, ['HORA', 'TICKETS TOTAL', 'INGRESOS TOTAL']); r3 += 1
    hora_map = {p['h']: p for p in por_hora_sem}
    max_t = max((p['tickets'] for p in por_hora_sem), default=0)
    for h in range(6, 22):
        p = hora_map.get(h, {'tickets':0,'total':0})
        fill = 'FFF3EE' if p['tickets']==max_t and max_t>0 else ('FFFFFF' if h%2==0 else 'F8FAFC')
        bar = '█' * min(int(p['tickets']/(max_t or 1)*10), 10) if p['tickets'] else ''
        cell(ws3, r3, 1, f'{h:02d}:00', align='center', fill_hex=fill)
        cell(ws3, r3, 2, f"{p['tickets']}  {bar}", fill_hex=fill)
        cell(ws3, r3, 3, p['total'], align='right', fill_hex=fill); ws3.cell(r3,3).number_format='$#,##0'
        r3 += 1

    # Hoja 4 — Cierres de caja
    ws4 = wb.create_sheet('💰 Cierres de caja')
    col_w(ws4, [20, 20, 16, 16, 16, 32])
    title_fn(ws4, 'Cierres de caja', semana_label)
    r4 = 3; hdr(ws4, r4, ['APERTURA', 'CIERRE', 'BASE', 'CIERRE REAL', 'DIFERENCIA', 'NOTAS']); r4 += 1
    for i, c in enumerate(cierres):
        fill = 'FEF2F2' if (c['difference'] or 0) < 0 else 'F0FDF4' if i%2==0 else 'FFFFFF'
        cell(ws4, r4, 1, str(c['opened_at'])[:16], align='center', fill_hex=fill)
        cell(ws4, r4, 2, str(c['closed_at'] or '—')[:16], align='center', fill_hex=fill)
        cell(ws4, r4, 3, c['opening_balance'] or 0, align='right', fill_hex=fill); ws4.cell(r4,3).number_format='$#,##0'
        cell(ws4, r4, 4, c['closing_balance'] or 0, align='right', fill_hex=fill); ws4.cell(r4,4).number_format='$#,##0'
        dif = c['difference'] or 0
        cell(ws4, r4, 5, dif, align='right', fill_hex=fill, bold=True, color=RED if dif<0 else GREEN); ws4.cell(r4,5).number_format='$#,##0'
        cell(ws4, r4, 6, c['notes'] or '—', fill_hex=fill)
        r4 += 1

    # Hoja 5 — Pérdidas
    ws5 = wb.create_sheet('📉 Pérdidas')
    col_w(ws5, [28, 16, 10, 16, 16, 28])
    title_fn(ws5, 'Pérdidas de la semana', semana_label)
    r5 = 3; hdr(ws5, r5, ['PRODUCTO', 'CATEGORÍA', 'CANT.', 'VALOR VENTA', 'RESPONSABLE', 'MOTIVO']); r5 += 1
    for i, p in enumerate(perdidas_sem):
        fill = 'FFFFFF' if i%2==0 else 'F8FAFC'
        cell(ws5, r5, 1, p['product_name'], fill_hex=fill)
        cell(ws5, r5, 2, p['category'] or '—', align='center', fill_hex=fill)
        cell(ws5, r5, 3, p['quantity'], align='center', fill_hex=fill)
        cell(ws5, r5, 4, p['sale_value'] or 0, align='right', fill_hex=fill, color=RED); ws5.cell(r5,4).number_format='$#,##0'
        cell(ws5, r5, 5, p['responsible'] or '—', align='center', fill_hex=fill)
        cell(ws5, r5, 6, p['reason'] or '—', fill_hex=fill); r5 += 1
    if perdidas_sem:
        hdr(ws5, r5, ['TOTAL', '', sum(p['quantity'] for p in perdidas_sem), sum(p['sale_value'] or 0 for p in perdidas_sem), '', ''], fill_hex='1E293B')
        ws5.cell(r5,4).number_format='$#,##0'

    # Hoja 6 — Pérdidas por trabajador
    ws6 = wb.create_sheet('👥 Trabajadores')
    col_w(ws6, [20, 14, 18])
    title_fn(ws6, 'Pérdidas por trabajador', semana_label)
    r6 = 3; hdr(ws6, r6, ['TRABAJADOR', 'N° PÉRDIDAS', 'VALOR TOTAL']); r6 += 1
    for i, w in enumerate(workers_perdidas):
        fill = 'FFFFFF' if i%2==0 else 'F8FAFC'
        cell(ws6, r6, 1, w['responsible'], bold=True, fill_hex=fill)
        cell(ws6, r6, 2, w['cnt'], align='center', fill_hex=fill)
        cell(ws6, r6, 3, w['total'], align='right', fill_hex=fill, color=RED); ws6.cell(r6,3).number_format='$#,##0'
        r6 += 1

    # Hoja 7 — Novedades
    ws7 = wb.create_sheet('📢 Novedades')
    col_w(ws7, [16, 42, 16, 18])
    title_fn(ws7, 'Novedades de la semana', semana_label)
    r7 = 3; hdr(ws7, r7, ['TIPO', 'DESCRIPCIÓN', 'REPORTADO POR', 'FECHA/HORA']); r7 += 1
    for i, n in enumerate(novedades_sem):
        fill = 'FFFFFF' if i%2==0 else 'F8FAFC'
        cell(ws7, r7, 1, n['tipo'].upper(), align='center', fill_hex=fill)
        cell(ws7, r7, 2, n['descripcion'], fill_hex=fill)
        cell(ws7, r7, 3, n['reportado_por'], align='center', fill_hex=fill)
        cell(ws7, r7, 4, str(n['created_at'])[:16], align='center', fill_hex=fill); r7 += 1
    if not novedades_sem: cell(ws7, r7, 1, 'Sin novedades', color='94A3B8')

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, as_attachment=True, download_name=f'informe_semanal_{d_ini}.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/api/reports/monthly', methods=['GET'])
def report_monthly():
    month = request.args.get('month', datetime.now().strftime('%Y-%m'))
    year, mon = int(month.split('-')[0]), int(month.split('-')[1])
    import calendar
    last_day = calendar.monthrange(year, mon)[1]
    d_ini = f'{month}-01'; d_fin = f'{month}-{last_day:02d}'
    month_label = datetime(year, mon, 1).strftime('%B %Y').capitalize()
    hdr, cell, title_fn, col_w, fmt_cop, ORANGE, DARK, GRAY, PURPLE, GREEN, RED, brd = _wb_style()
    wb = openpyxl.Workbook()

    with get_db() as conn:
        dias = conn.execute("""SELECT date(created_at) dia, COUNT(*) tickets, COALESCE(SUM(total),0) total
            FROM sales WHERE date(created_at) BETWEEN ? AND ? GROUP BY dia ORDER BY dia""", (d_ini, d_fin)).fetchall()
        egr_mes = conn.execute("""SELECT date(created_at) dia, COALESCE(SUM(amount),0) t
            FROM cash_movements WHERE type='egreso' AND date(created_at) BETWEEN ? AND ? GROUP BY dia""", (d_ini, d_fin)).fetchall()
        ing_mes = conn.execute("""SELECT date(created_at) dia, COALESCE(SUM(amount),0) t
            FROM cash_movements WHERE type='ingreso' AND date(created_at) BETWEEN ? AND ? GROUP BY dia""", (d_ini, d_fin)).fetchall()
        top_prods = conn.execute("""SELECT si.product_name, SUM(si.quantity) qty, SUM(si.subtotal) total
            FROM sale_items si JOIN sales s ON s.id=si.sale_id
            WHERE date(s.created_at) BETWEEN ? AND ? GROUP BY si.product_name ORDER BY total DESC""", (d_ini, d_fin)).fetchall()
        por_hora = conn.execute("""SELECT CAST(strftime('%H', created_at) AS INTEGER) h,
            COUNT(*) tickets, COALESCE(SUM(total),0) total
            FROM sales WHERE date(created_at) BETWEEN ? AND ? GROUP BY h ORDER BY h""", (d_ini, d_fin)).fetchall()
        cierres = conn.execute("""SELECT opened_at, closing_balance, difference, notes
            FROM cash_registers WHERE date(opened_at) BETWEEN ? AND ? ORDER BY opened_at""", (d_ini, d_fin)).fetchall()
        perdidas = conn.execute("""SELECT product_name, category, quantity, sale_value, responsible, reason, created_at
            FROM losses WHERE date(created_at) BETWEEN ? AND ? ORDER BY created_at""", (d_ini, d_fin)).fetchall()
        novedades = conn.execute("""SELECT tipo, descripcion, reportado_por, created_at
            FROM novedades WHERE date(created_at) BETWEEN ? AND ? ORDER BY created_at""", (d_ini, d_fin)).fetchall()
        workers_stats = conn.execute("""SELECT responsible, COUNT(*) cnt, COALESCE(SUM(sale_value),0) total
            FROM losses WHERE date(created_at) BETWEEN ? AND ? AND responsible IS NOT NULL AND responsible!=''
            GROUP BY responsible ORDER BY total DESC""", (d_ini, d_fin)).fetchall()
        descuadres = conn.execute("""SELECT difference, opened_at FROM cash_registers
            WHERE date(opened_at) BETWEEN ? AND ? AND difference < 0 ORDER BY opened_at""", (d_ini, d_fin)).fetchall()
        semanas_data = []
        from datetime import timedelta as td
        ref = datetime(year, mon, 1)
        sem_start = ref
        while sem_start.month == mon:
            sem_end = min(sem_start + td(days=6), datetime(year, mon, last_day))
            row_s = conn.execute("""SELECT COUNT(*) tickets, COALESCE(SUM(total),0) total FROM sales
                WHERE date(created_at) BETWEEN ? AND ?""",
                (sem_start.strftime('%Y-%m-%d'), sem_end.strftime('%Y-%m-%d'))).fetchone()
            semanas_data.append({'label': f'{sem_start.strftime("%d/%m")}–{sem_end.strftime("%d/%m")}',
                                 'tickets': row_s['tickets'], 'total': row_s['total']})
            sem_start = sem_start + td(days=7)

    # Hoja 1 — Resumen mensual
    ws = wb.active; ws.title = '📊 Resumen Mensual'
    col_w(ws, [16, 12, 18, 16, 16, 16, 12])
    title_fn(ws, f'Informe Mensual — {month_label}', f'Periodo: {d_ini} → {d_fin}  |  Generado: {datetime.now().strftime("%Y-%m-%d %H:%M")}')
    r = 3
    dias_nombres = {'Monday':'Lun','Tuesday':'Mar','Wednesday':'Mié','Thursday':'Jue','Friday':'Vie','Saturday':'Sáb','Sunday':'Dom'}
    hdr(ws, r, ['FECHA', 'DÍA', 'TICKETS', 'VENTAS', 'INGRESOS', 'EGRESOS', 'BALANCE']); r += 1
    dia_map = {d['dia']: d for d in dias}
    egr_map = {e['dia']: e['t'] for e in egr_mes}; ing_map = {e['dia']: e['t'] for e in ing_mes}
    t_tick=t_ven=t_ing=t_egr = 0
    from datetime import date as dt_date
    for day in range(1, last_day+1):
        d = f'{month}-{day:02d}'
        dd = datetime(year, mon, day)
        dn = dias_nombres.get(dd.strftime('%A'), '')
        di = dia_map.get(d, {'tickets':0,'total':0})
        ing = ing_map.get(d,0); egr = egr_map.get(d,0); bal = ing-egr
        is_weekend = dd.weekday() >= 5
        fill = 'EFF6FF' if is_weekend else ('FFFFFF' if day%2==0 else 'F8FAFC')
        cell(ws, r, 1, d[8:]+'/'+d[5:7], align='center', fill_hex=fill)
        cell(ws, r, 2, dn, align='center', fill_hex=fill, color='6366F1' if is_weekend else DARK)
        cell(ws, r, 3, di['tickets'], align='center', fill_hex=fill)
        cell(ws, r, 4, di['total'], align='right', fill_hex=fill); ws.cell(r,4).number_format='$#,##0'
        cell(ws, r, 5, ing, align='right', fill_hex=fill); ws.cell(r,5).number_format='$#,##0'
        cell(ws, r, 6, egr, align='right', fill_hex=fill, color=RED); ws.cell(r,6).number_format='$#,##0'
        cell(ws, r, 7, bal, align='right', fill_hex=fill, color=GREEN if bal>=0 else RED, bold=True); ws.cell(r,7).number_format='$#,##0'
        t_tick+=di['tickets']; t_ven+=di['total']; t_ing+=ing; t_egr+=egr; r+=1
    hdr(ws, r, ['TOTALES', '', t_tick, t_ven, t_ing, t_egr, t_ing-t_egr], fill_hex='1E293B')
    for c in [4,5,6,7]: ws.cell(r,c).number_format='$#,##0'

    # Hoja 2 — Comparación semanas
    ws2 = wb.create_sheet('📅 Por Semana')
    col_w(ws2, [22, 14, 18, 16])
    title_fn(ws2, 'Comparación por semana', month_label)
    r2 = 3; hdr(ws2, r2, ['SEMANA', 'TICKETS', 'VENTAS', 'PROMEDIO/DÍA']); r2 += 1
    for i, s in enumerate(semanas_data):
        fill = 'FFFFFF' if i%2==0 else 'F8FAFC'
        cell(ws2, r2, 1, f'Semana {i+1}  ({s["label"]})', fill_hex=fill)
        cell(ws2, r2, 2, s['tickets'], align='center', fill_hex=fill)
        cell(ws2, r2, 3, s['total'], align='right', fill_hex=fill); ws2.cell(r2,3).number_format='$#,##0'
        cell(ws2, r2, 4, round(s['total']/7), align='right', fill_hex=fill); ws2.cell(r2,4).number_format='$#,##0'
        r2 += 1

    # Hoja 3 — Top productos
    ws3 = wb.create_sheet('🛒 Top Productos')
    col_w(ws3, [34, 12, 18, 14, 12])
    title_fn(ws3, 'Productos más vendidos', month_label)
    r3 = 3; hdr(ws3, r3, ['PRODUCTO', 'CANTIDAD', 'TOTAL', '% PARTICIPACIÓN', 'RANKING']); r3 += 1
    tot_p = sum(p['total'] for p in top_prods) or 1
    for i, p in enumerate(top_prods):
        fill = 'FFF3EE' if i==0 else 'FFFFFF' if i%2==0 else 'F8FAFC'
        medal = ['🥇','🥈','🥉']
        cell(ws3, r3, 1, p['product_name'], bold=(i==0), fill_hex=fill)
        cell(ws3, r3, 2, int(p['qty']), align='center', fill_hex=fill)
        cell(ws3, r3, 3, p['total'], align='right', fill_hex=fill); ws3.cell(r3,3).number_format='$#,##0'
        cell(ws3, r3, 4, f'{round(p["total"]/tot_p*100,1)}%', align='center', fill_hex=fill)
        cell(ws3, r3, 5, medal[i] if i<3 else str(i+1), align='center', fill_hex=fill); r3 += 1

    # Hoja 4 — Horas pico
    ws4 = wb.create_sheet('⏰ Horas Pico')
    col_w(ws4, [14, 14, 18])
    title_fn(ws4, 'Horas pico del mes', month_label)
    r4 = 3; hdr(ws4, r4, ['HORA', 'TICKETS', 'INGRESOS']); r4 += 1
    h_map = {p['h']: p for p in por_hora}
    max_h = max((p['tickets'] for p in por_hora), default=0)
    for h in range(6, 22):
        p = h_map.get(h, {'tickets':0,'total':0})
        fill = 'FFF3EE' if p['tickets']==max_h and max_h>0 else ('FFFFFF' if h%2==0 else 'F8FAFC')
        bar = '█'*min(int(p['tickets']/(max_h or 1)*10),10) if p['tickets'] else ''
        cell(ws4, r4, 1, f'{h:02d}:00', align='center', fill_hex=fill)
        cell(ws4, r4, 2, f"{p['tickets']}  {bar}", fill_hex=fill)
        cell(ws4, r4, 3, p['total'], align='right', fill_hex=fill); ws4.cell(r4,3).number_format='$#,##0'
        r4 += 1

    # Hoja 5 — Descuadres de caja
    ws5 = wb.create_sheet('💰 Caja y Descuadres')
    col_w(ws5, [22, 18, 16])
    title_fn(ws5, 'Descuadres de caja', month_label)
    r5 = 3; hdr(ws5, r5, ['FECHA', 'CIERRE REAL', 'DIFERENCIA']); r5 += 1
    total_descuadre = 0
    for i, c in enumerate(descuadres):
        fill = 'FEF2F2'
        cell(ws5, r5, 1, str(c['opened_at'])[:10], align='center', fill_hex=fill)
        cell(ws5, r5, 2, c['closing_balance'] or 0, align='right', fill_hex=fill); ws5.cell(r5,2).number_format='$#,##0'
        cell(ws5, r5, 3, c['difference'] or 0, align='right', fill_hex=fill, color=RED, bold=True); ws5.cell(r5,3).number_format='$#,##0'
        total_descuadre += (c['difference'] or 0); r5 += 1
    if not descuadres: cell(ws5, r5, 1, '✅ Sin descuadres este mes', color=GREEN, bold=True)
    else:
        r5 += 1; cell(ws5, r5, 1, 'TOTAL DESCUADRES', bold=True); cell(ws5, r5, 3, total_descuadre, color=RED, bold=True); ws5.cell(r5,3).number_format='$#,##0'

    # Hoja 6 — Pérdidas
    ws6 = wb.create_sheet('📉 Pérdidas')
    col_w(ws6, [28, 16, 10, 16, 16, 28])
    title_fn(ws6, 'Pérdidas del mes', month_label)
    r6 = 3; hdr(ws6, r6, ['PRODUCTO', 'CATEGORÍA', 'CANT.', 'VALOR', 'RESPONSABLE', 'MOTIVO']); r6 += 1
    for i, p in enumerate(perdidas):
        fill = 'FFFFFF' if i%2==0 else 'F8FAFC'
        cell(ws6, r6, 1, p['product_name'], fill_hex=fill); cell(ws6, r6, 2, p['category'] or '—', align='center', fill_hex=fill)
        cell(ws6, r6, 3, p['quantity'], align='center', fill_hex=fill)
        cell(ws6, r6, 4, p['sale_value'] or 0, align='right', fill_hex=fill, color=RED); ws6.cell(r6,4).number_format='$#,##0'
        cell(ws6, r6, 5, p['responsible'] or '—', align='center', fill_hex=fill)
        cell(ws6, r6, 6, p['reason'] or '—', fill_hex=fill); r6 += 1

    # Hoja 7 — Trabajadores
    ws7 = wb.create_sheet('👥 Trabajadores')
    col_w(ws7, [20, 14, 18])
    title_fn(ws7, 'Pérdidas por trabajador', month_label)
    r7 = 3; hdr(ws7, r7, ['TRABAJADOR', 'N° PÉRDIDAS', 'VALOR TOTAL']); r7 += 1
    for i, w in enumerate(workers_stats):
        fill = 'FFFFFF' if i%2==0 else 'F8FAFC'
        cell(ws7, r7, 1, w['responsible'], bold=True, fill_hex=fill)
        cell(ws7, r7, 2, w['cnt'], align='center', fill_hex=fill)
        cell(ws7, r7, 3, w['total'], align='right', fill_hex=fill, color=RED); ws7.cell(r7,3).number_format='$#,##0'; r7 += 1

    # Hoja 8 — Novedades
    ws8 = wb.create_sheet('📢 Novedades')
    col_w(ws8, [16, 42, 16, 18])
    title_fn(ws8, 'Novedades del mes', month_label)
    r8 = 3; hdr(ws8, r8, ['TIPO', 'DESCRIPCIÓN', 'REPORTADO POR', 'FECHA']); r8 += 1
    for i, n in enumerate(novedades):
        fill = 'FFFFFF' if i%2==0 else 'F8FAFC'
        cell(ws8, r8, 1, n['tipo'].upper(), align='center', fill_hex=fill)
        cell(ws8, r8, 2, n['descripcion'], fill_hex=fill)
        cell(ws8, r8, 3, n['reportado_por'], align='center', fill_hex=fill)
        cell(ws8, r8, 4, str(n['created_at'])[:16], align='center', fill_hex=fill); r8 += 1
    if not novedades: cell(ws8, r8, 1, 'Sin novedades', color='94A3B8')

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, as_attachment=True, download_name=f'informe_mensual_{month}.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ─── CIERRES DE CAJA ─────────────────────────────────
@app.route('/api/registers', methods=['GET'])
def get_registers():
    with get_db() as conn:
        rows = conn.execute('SELECT * FROM cash_registers ORDER BY opened_at DESC LIMIT 30').fetchall()
    return jsonify(rows_to_list(rows))

@app.route('/api/registers/current', methods=['GET'])
def get_current_register():
    with get_db() as conn:
        row = conn.execute("SELECT * FROM cash_registers WHERE status='abierta' ORDER BY opened_at DESC LIMIT 1").fetchone()
    return jsonify(row_to_dict(row))

@app.route('/api/registers', methods=['POST'])
def open_register():
    d = request.json or {}
    with get_db() as conn:
        # Solo puede haber una caja abierta
        existing = conn.execute("SELECT id FROM cash_registers WHERE status='abierta'").fetchone()
        if existing:
            return jsonify({'error': 'Ya hay una caja abierta'}), 400
        cur = conn.execute(
            'INSERT INTO cash_registers (opening_balance, notes) VALUES (?,?)',
            (d.get('opening_balance', 0), d.get('notes', ''))
        )
        row = conn.execute('SELECT * FROM cash_registers WHERE id=?', (cur.lastrowid,)).fetchone()
    return jsonify(row_to_dict(row))

@app.route('/api/registers/current/summary', methods=['GET'])
def current_register_summary():
    with get_db() as conn:
        reg = conn.execute("SELECT * FROM cash_registers WHERE status='abierta' ORDER BY opened_at DESC LIMIT 1").fetchone()
        if not reg:
            return jsonify({'error': 'No hay caja abierta'}), 404
        opened_at = reg['opened_at']
        sales_by_method_rows = conn.execute(
            "SELECT payment_method, COALESCE(SUM(total),0) as total, COUNT(*) as count FROM sales WHERE created_at >= ? GROUP BY payment_method",
            (opened_at,)
        ).fetchall()
        mov_in = conn.execute(
            "SELECT COALESCE(SUM(amount),0) as t FROM cash_movements WHERE type='ingreso' AND description NOT LIKE 'Venta #%' AND created_at >= ?",
            (opened_at,)
        ).fetchone()
        mov_out = conn.execute(
            "SELECT COALESCE(SUM(amount),0) as t FROM cash_movements WHERE type='egreso' AND created_at >= ?",
            (opened_at,)
        ).fetchone()
        total_orders = conn.execute(
            "SELECT COUNT(*) as c, COALESCE(SUM(total),0) as t FROM sales WHERE created_at >= ?",
            (opened_at,)
        ).fetchone()
    sbm = {r['payment_method']: {'total': r['total'], 'count': r['count']} for r in sales_by_method_rows}
    return jsonify({
        'opening_balance': reg['opening_balance'],
        'opened_at': reg['opened_at'],
        'sales_by_method': sbm,
        'manual_in': mov_in['t'],
        'manual_out': mov_out['t'],
        'total_orders': total_orders['c'],
        'total_amount': total_orders['t'],
    })

@app.route('/api/registers/<int:rid>/close', methods=['PUT'])
def close_register(rid):
    d = request.json or {}
    counted_cash = d.get('counted_cash', 0)
    notes = d.get('notes', '')
    with get_db() as conn:
        reg = conn.execute('SELECT * FROM cash_registers WHERE id=?', (rid,)).fetchone()
        if not reg or reg['status'] == 'cerrada':
            return jsonify({'error': 'Caja no encontrada o ya cerrada'}), 400

        opened_at = reg['opened_at']
        # Calcular totales desde apertura
        sales = conn.execute(
            "SELECT COALESCE(SUM(total),0) as t, COALESCE(SUM(CASE WHEN payment_method='efectivo' THEN total ELSE 0 END),0) as cash_t FROM sales WHERE created_at >= ?",
            (opened_at,)
        ).fetchone()
        mov_in = conn.execute(
            "SELECT COALESCE(SUM(amount),0) as t FROM cash_movements WHERE type='ingreso' AND description NOT LIKE 'Venta #%' AND created_at >= ?",
            (opened_at,)
        ).fetchone()
        mov_out = conn.execute(
            "SELECT COALESCE(SUM(amount),0) as t FROM cash_movements WHERE type='egreso' AND created_at >= ?",
            (opened_at,)
        ).fetchone()

        total_sales = sales['t']
        total_cash_sales = sales['cash_t']
        total_in = mov_in['t']
        total_out = mov_out['t']
        expected_cash = reg['opening_balance'] + total_cash_sales + total_in - total_out
        difference = counted_cash - expected_cash

        # Ventas agrupadas por método de pago
        sales_by_method_rows = conn.execute(
            "SELECT payment_method, COALESCE(SUM(total),0) as total FROM sales WHERE created_at >= ? GROUP BY payment_method",
            (opened_at,)
        ).fetchall()
        sales_by_method = {r['payment_method']: r['total'] for r in sales_by_method_rows}

        conn.execute("""
            UPDATE cash_registers SET
                status='cerrada', counted_cash=?, total_sales=?, total_cash_sales=?,
                total_in=?, total_out=?, expected_cash=?, difference=?, notes=?,
                closed_at=datetime('now','localtime')
            WHERE id=?
        """, (counted_cash, total_sales, total_cash_sales, total_in, total_out,
              expected_cash, difference, notes, rid))

        # Registrar el efectivo contado como egreso de cierre
        closed_date = datetime.now().strftime('%d/%m/%Y')
        turno_label = notes.split(' — ')[0] if notes else 'Turno'
        cierre_desc = f"CIERRE DE CAJA {closed_date} — {turno_label}"
        conn.execute(
            "INSERT INTO cash_movements (type, amount, description, category) VALUES ('egreso', ?, ?, 'Cierre de caja')",
            (counted_cash, cierre_desc)
        )
        # Si hay diferencia, registrarla también
        if difference != 0:
            diff_desc = f"Diferencia de efectivo — {closed_date}"
            diff_type = 'ingreso' if difference > 0 else 'egreso'
            conn.execute(
                "INSERT INTO cash_movements (type, amount, description, category) VALUES (?, ?, ?, 'Cierre de caja')",
                (diff_type, abs(difference), diff_desc)
            )

        # ── Asientos contables automáticos del cierre ──
        today_str = datetime.now().strftime('%Y-%m-%d')
        ref_cierre = f'Cierre {closed_date} — {turno_label}'

        acc_ventas = conn.execute("SELECT id FROM accounts WHERE code='4001'").fetchone()
        if acc_ventas and total_sales > 0:
            conn.execute("""
                INSERT INTO journal_entries (date, account_id, description, entry_type, amount, reference)
                VALUES (?, ?, ?, 'ingreso', ?, ?)
            """, (today_str, acc_ventas['id'],
                  f'Ventas del turno — {turno_label}', total_sales, ref_cierre))

        if total_out > 0:
            acc_egresos = conn.execute("SELECT id FROM accounts WHERE code='8004'").fetchone()
            if acc_egresos:
                conn.execute("""
                    INSERT INTO journal_entries (date, account_id, description, entry_type, amount, reference)
                    VALUES (?, ?, ?, 'egreso', ?, ?)
                """, (today_str, acc_egresos['id'],
                      f'Egresos del turno — {turno_label}', total_out, ref_cierre))

        # Pérdidas del turno
        losses_total = conn.execute(
            "SELECT COALESCE(SUM(quantity * 1),0) as cnt, COUNT(*) as c FROM losses WHERE created_at >= ?",
            (opened_at,)
        ).fetchone()

        row = conn.execute('SELECT * FROM cash_registers WHERE id=?', (rid,)).fetchone()

        # Configuración WhatsApp
        cfg = {r['key']: r['value'] for r in conn.execute('SELECT key,value FROM settings').fetchall()}

    result = row_to_dict(row)
    result['sales_by_method'] = sales_by_method
    result['total_in'] = total_in
    result['total_out'] = total_out

    # Enviar WhatsApp via Green API si está habilitado
    if cfg.get('greenapi_enabled') == '1' and cfg.get('whatsapp_phone') and cfg.get('greenapi_instance') and cfg.get('greenapi_token'):
        nombre = cfg.get('negocio_nombre', 'Cafetería')
        fecha = datetime.now().strftime('%d/%m/%Y %H:%M')
        diff_sign = '+' if difference >= 0 else ''
        metodos_str = ' | '.join([f'{k}: ${v:,.0f}' for k, v in sales_by_method.items()]) or 'Sin ventas'
        msg = (
            f'🏪 {nombre} — Cierre de caja\n'
            f'📅 {fecha}\n\n'
            f'💰 Ventas totales: ${total_sales:,.0f}\n'
            f'   {metodos_str}\n'
            f'💸 Egresos: ${total_out:,.0f}\n\n'
            f'📦 Fondo inicial: ${result["opening_balance"]:,.0f}\n'
            f'✅ Efectivo esperado: ${result["expected_cash"]:,.0f}\n'
            f'🔢 Efectivo contado: ${counted_cash:,.0f}\n'
            f'{"🟢" if difference >= 0 else "🔴"} Diferencia: {diff_sign}${difference:,.0f}'
        )
        import threading
        threading.Thread(target=send_whatsapp, args=(cfg['whatsapp_phone'], cfg['greenapi_instance'], cfg['greenapi_token'], msg), daemon=True).start()

    return jsonify(result)

# ─── COMANDAS ────────────────────────────────────────
@app.route('/api/comandas', methods=['GET'])
def get_comandas():
    status = request.args.get('status')
    with get_db() as conn:
        if status:
            rows = conn.execute(
                "SELECT * FROM comandas WHERE status=? ORDER BY created_at DESC", (status,)
            ).fetchall()
        else:
            rows = conn.execute(
                "SELECT * FROM comandas ORDER BY created_at DESC LIMIT 50"
            ).fetchall()
    result = []
    for r in rows:
        d = dict(r)
        d['items'] = json.loads(d['items'])
        result.append(d)
    return jsonify(result)

@app.route('/api/comandas', methods=['POST'])
def create_comanda():
    d = request.json
    items = d.get('items', [])
    if not items:
        return jsonify({'error': 'La comanda no tiene productos'}), 400
    with get_db() as conn:
        cur = conn.execute(
            "INSERT INTO comandas (customer_name, items, notes) VALUES (?,?,?)",
            (d.get('customer_name',''), json.dumps(items), d.get('notes',''))
        )
        row = conn.execute('SELECT * FROM comandas WHERE id=?', (cur.lastrowid,)).fetchone()
    result = dict(row)
    result['items'] = json.loads(result['items'])
    return jsonify(result)

@app.route('/api/comandas/<int:cid>/status', methods=['PUT'])
def update_comanda_status(cid):
    d = request.json
    new_status = d.get('status')
    if new_status not in ('pendiente','listo','entregado','cancelado'):
        return jsonify({'error': 'Estado inválido'}), 400
    with get_db() as conn:
        conn.execute('UPDATE comandas SET status=? WHERE id=?', (new_status, cid))
        row = conn.execute('SELECT * FROM comandas WHERE id=?', (cid,)).fetchone()
    result = dict(row)
    result['items'] = json.loads(result['items'])
    return jsonify(result)

# ─── DEVOLUCIONES ────────────────────────────────────
@app.route('/api/returns', methods=['GET'])
def get_returns():
    with get_db() as conn:
        rows = conn.execute('SELECT * FROM returns ORDER BY created_at DESC LIMIT 50').fetchall()
    return jsonify(rows_to_list(rows))

@app.route('/api/returns', methods=['POST'])
def create_return():
    d = request.json or {}
    sale_id = d.get('sale_id')
    items = d.get('items', [])
    reason = d.get('reason', '')
    if not sale_id or not items:
        return jsonify({'error': 'Venta e ítems requeridos'}), 400
    with get_db() as conn:
        sale = conn.execute('SELECT * FROM sales WHERE id=?', (sale_id,)).fetchone()
        if not sale:
            return jsonify({'error': 'Venta no encontrada'}), 404
        total = sum(i['price'] * i['quantity'] for i in items)
        cur = conn.execute(
            'INSERT INTO returns (sale_id, total, reason, payment_method) VALUES (?,?,?,?)',
            (sale_id, total, reason, sale['payment_method'])
        )
        ret_id = cur.lastrowid
        for item in items:
            conn.execute(
                'INSERT INTO return_items (return_id, product_id, product_name, price, quantity, subtotal) VALUES (?,?,?,?,?,?)',
                (ret_id, item.get('product_id'), item['product_name'],
                 item['price'], item['quantity'], item['price'] * item['quantity'])
            )
            # Restaurar stock
            if item.get('product_id'):
                conn.execute('UPDATE products SET stock = stock + ? WHERE id=?',
                             (item['quantity'], item['product_id']))
        # Registrar egreso (devolución de dinero al cliente)
        conn.execute(
            "INSERT INTO cash_movements (type, amount, description, category) VALUES ('egreso',?,?,?)",
            (total, f'Devolución — Venta #{sale_id}', 'Devoluciones')
        )
        row = conn.execute('SELECT * FROM returns WHERE id=?', (ret_id,)).fetchone()
    return jsonify(row_to_dict(row))

# ─── CONTABILIDAD ─────────────────────────────────────
@app.route('/api/accounts', methods=['GET'])
def get_accounts():
    with get_db() as conn:
        rows = conn.execute('SELECT * FROM accounts WHERE active=1 ORDER BY code').fetchall()
    return jsonify(rows_to_list(rows))

@app.route('/api/accounts', methods=['POST'])
def create_account():
    d = request.json or {}
    if not d.get('name') or not d.get('code') or not d.get('type'):
        return jsonify({'error': 'Código, nombre y tipo requeridos'}), 400
    with get_db() as conn:
        cur = conn.execute('INSERT INTO accounts (code,name,type) VALUES (?,?,?)',
                           (d['code'], d['name'], d['type']))
        row = conn.execute('SELECT * FROM accounts WHERE id=?', (cur.lastrowid,)).fetchone()
    return jsonify(row_to_dict(row))

@app.route('/api/journal', methods=['GET'])
def get_journal():
    from_d = request.args.get('from')
    to_d = request.args.get('to')
    limit = int(request.args.get('limit', 200))
    with get_db() as conn:
        if from_d and to_d:
            rows = conn.execute("""
                SELECT j.*, a.code, a.name as account_name, a.type as account_type
                FROM journal_entries j JOIN accounts a ON a.id=j.account_id
                WHERE date(j.date) BETWEEN date(?) AND date(?) ORDER BY j.date DESC, j.id DESC LIMIT ?
            """, (from_d, to_d, limit)).fetchall()
        else:
            rows = conn.execute("""
                SELECT j.*, a.code, a.name as account_name, a.type as account_type
                FROM journal_entries j JOIN accounts a ON a.id=j.account_id
                ORDER BY j.date DESC, j.id DESC LIMIT ?
            """, (limit,)).fetchall()
    return jsonify(rows_to_list(rows))

@app.route('/api/journal', methods=['POST'])
def create_journal_entry():
    d = request.json or {}
    if not d.get('account_id') or not d.get('amount') or not d.get('description') or not d.get('date'):
        return jsonify({'error': 'Cuenta, monto, descripción y fecha requeridos'}), 400
    with get_db() as conn:
        acc = conn.execute('SELECT * FROM accounts WHERE id=?', (d['account_id'],)).fetchone()
        if not acc:
            return jsonify({'error': 'Cuenta no encontrada'}), 404
        entry_type = 'ingreso' if acc['type'] in ('ingreso','otro_ingreso') else 'egreso'
        cur = conn.execute(
            'INSERT INTO journal_entries (date,account_id,description,entry_type,amount,reference) VALUES (?,?,?,?,?,?)',
            (d['date'], d['account_id'], d['description'], entry_type, d['amount'], d.get('reference',''))
        )
        row = conn.execute("""
            SELECT j.*, a.code, a.name as account_name, a.type as account_type
            FROM journal_entries j JOIN accounts a ON a.id=j.account_id WHERE j.id=?
        """, (cur.lastrowid,)).fetchone()
    return jsonify(row_to_dict(row))

@app.route('/api/journal/<int:jid>', methods=['DELETE'])
def delete_journal_entry(jid):
    with get_db() as conn:
        conn.execute('DELETE FROM journal_entries WHERE id=?', (jid,))
    return jsonify({'ok': True})

@app.route('/api/ledger', methods=['GET'])
def get_ledger():
    from_d = request.args.get('from')
    to_d = request.args.get('to')
    account_id = request.args.get('account_id')
    with get_db() as conn:
        accounts = conn.execute('SELECT * FROM accounts WHERE active=1 ORDER BY code').fetchall()
        result = []
        for acc in accounts:
            if account_id and str(acc['id']) != str(account_id):
                continue
            params = [acc['id']]
            where = 'account_id=?'
            if from_d and to_d:
                where += ' AND date(date) BETWEEN date(?) AND date(?)'
                params += [from_d, to_d]
            entries = conn.execute(
                f'SELECT * FROM journal_entries WHERE {where} ORDER BY date ASC, id ASC', params
            ).fetchall()
            total_in  = sum(e['amount'] for e in entries if e['entry_type']=='ingreso')
            total_out = sum(e['amount'] for e in entries if e['entry_type']=='egreso')
            result.append({
                'account': dict(acc),
                'entries': rows_to_list(entries),
                'total_ingreso': total_in,
                'total_egreso': total_out,
                'saldo': total_in - total_out
            })
    return jsonify(result)

@app.route('/api/acc-ventas-dia', methods=['GET'])
def get_acc_ventas_dia():
    from_d = request.args.get('from')
    to_d   = request.args.get('to')
    with get_db() as conn:
        if from_d and to_d:
            rows = conn.execute("""
                SELECT
                    date(created_at) as dia,
                    COUNT(*) as num_ventas,
                    COALESCE(SUM(total),0) as total,
                    COALESCE(SUM(CASE WHEN payment_method='efectivo' THEN total ELSE 0 END),0) as efectivo,
                    COALESCE(SUM(CASE WHEN payment_method='transferencia' THEN total ELSE 0 END),0) as transferencia
                FROM sales
                WHERE date(created_at) BETWEEN date(?) AND date(?)
                GROUP BY date(created_at)
                ORDER BY dia DESC
            """, (from_d, to_d)).fetchall()
        else:
            rows = conn.execute("""
                SELECT
                    date(created_at) as dia,
                    COUNT(*) as num_ventas,
                    COALESCE(SUM(total),0) as total,
                    COALESCE(SUM(CASE WHEN payment_method='efectivo' THEN total ELSE 0 END),0) as efectivo,
                    COALESCE(SUM(CASE WHEN payment_method='transferencia' THEN total ELSE 0 END),0) as transferencia
                FROM sales
                GROUP BY date(created_at)
                ORDER BY dia DESC
                LIMIT 60
            """).fetchall()
        result = rows_to_list(rows)
    return jsonify({
        'rows': result,
        'total_ventas': sum(r['num_ventas'] for r in result),
        'total_monto': sum(r['total'] for r in result),
        'total_efectivo': sum(r['efectivo'] for r in result),
        'total_transferencia': sum(r['transferencia'] for r in result),
    })

@app.route('/api/available-months', methods=['GET'])
def get_available_months():
    from datetime import datetime
    meses = ['Enero','Febrero','Marzo','Abril','Mayo','Junio',
             'Julio','Agosto','Septiembre','Octubre','Noviembre','Diciembre']
    now = datetime.now()
    cur_year = now.year
    cur_month = now.month

    # Collect all months with data
    with get_db() as conn:
        rows = conn.execute("""
            SELECT strftime('%Y-%m', date) as ym FROM journal_entries GROUP BY ym
            UNION
            SELECT strftime('%Y-%m', created_at) as ym FROM sales GROUP BY ym
        """).fetchall()
    data_months = {r['ym'] for r in rows}

    # All 12 months of current year + any months with data from other years
    all_yms = set()
    for m in range(1, 13):
        all_yms.add(f"{cur_year}-{str(m).zfill(2)}")
    # Add past-year months that have data
    for ym in data_months:
        y = int(ym.split('-')[0])
        if y != cur_year:
            all_yms.add(ym)

    months = []
    for ym in sorted(all_yms, reverse=True):
        y, m = ym.split('-')
        months.append({'ym': ym, 'year': int(y), 'month': int(m),
                       'label': f"{meses[int(m)-1]} {y}"})
    return jsonify(months)

@app.route('/api/acc-summary', methods=['GET'])
def get_acc_summary():
    from_d = request.args.get('from')
    to_d   = request.args.get('to')
    with get_db() as conn:
        if from_d and to_d:
            sales_row = conn.execute(
                "SELECT COUNT(*) as count, COALESCE(SUM(total),0) as total FROM sales WHERE date(created_at) BETWEEN date(?) AND date(?)",
                (from_d, to_d)
            ).fetchone()
            egresos_row = conn.execute(
                "SELECT COALESCE(SUM(amount),0) as total FROM cash_movements WHERE type='egreso' AND date(created_at) BETWEEN date(?) AND date(?)",
                (from_d, to_d)
            ).fetchone()
            ingresos_mov = conn.execute(
                "SELECT COALESCE(SUM(amount),0) as total FROM cash_movements WHERE type='ingreso' AND date(created_at) BETWEEN date(?) AND date(?)",
                (from_d, to_d)
            ).fetchone()
        else:
            sales_row = conn.execute(
                "SELECT COUNT(*) as count, COALESCE(SUM(total),0) as total FROM sales"
            ).fetchone()
            egresos_row = conn.execute(
                "SELECT COALESCE(SUM(amount),0) as total FROM cash_movements WHERE type='egreso'"
            ).fetchone()
            ingresos_mov = conn.execute(
                "SELECT COALESCE(SUM(amount),0) as total FROM cash_movements WHERE type='ingreso'"
            ).fetchone()
        # Dinero en caja = total ingresos movimientos - total egresos (sin filtro de fecha, saldo acumulado)
        caja_ing = conn.execute("SELECT COALESCE(SUM(amount),0) as t FROM cash_movements WHERE type='ingreso'").fetchone()
        caja_egr = conn.execute("SELECT COALESCE(SUM(amount),0) as t FROM cash_movements WHERE type='egreso'").fetchone()
    return jsonify({
        'total_ventas': sales_row['total'],
        'num_ventas': sales_row['count'],
        'total_egresos': egresos_row['total'],
        'total_ingresos_mov': ingresos_mov['total'],
        'dinero_en_caja': caja_ing['t'] - caja_egr['t'],
    })

@app.route('/api/income-statement', methods=['GET'])
def get_income_statement():
    from_d = request.args.get('from')
    to_d = request.args.get('to')
    with get_db() as conn:
        params_base = []
        date_filter = ''
        if from_d and to_d:
            date_filter = ' AND date(j.date) BETWEEN date(?) AND date(?)'
            params_base = [from_d, to_d]

        def get_by_type(acc_type, code_prefix=None):
            code_cond = f" AND a.code LIKE '{code_prefix}%'" if code_prefix else ''
            rows = conn.execute(f"""
                SELECT a.id, a.code, a.name, COALESCE(SUM(j.amount),0) as total
                FROM accounts a
                LEFT JOIN journal_entries j ON j.account_id=a.id{date_filter}
                WHERE a.type=? AND a.active=1{code_cond}
                GROUP BY a.id ORDER BY a.code
            """, params_base + [acc_type]).fetchall()
            return rows_to_list(rows)

        ingresos        = get_by_type('ingreso')
        costos          = get_by_type('costo')
        gastos_op       = get_by_type('gasto_operativo')
        gastos_fin      = get_by_type('gasto_admin')        # 7xxx — Gastos Financieros
        otros_gastos    = get_by_type('otro_gasto', '8')    # 8xxx — Otros Gastos
        impuestos_list  = get_by_type('otro_gasto', '9')    # 9xxx — Impuestos

        total_ingresos  = sum(r['total'] for r in ingresos)
        total_costos    = sum(r['total'] for r in costos)
        total_gastos_op = sum(r['total'] for r in gastos_op)
        total_gastos_fin= sum(r['total'] for r in gastos_fin)
        total_otros     = sum(r['total'] for r in otros_gastos)
        total_impuestos = sum(r['total'] for r in impuestos_list)

        ganancia_bruta          = total_ingresos - total_costos
        pct_bruta               = round(ganancia_bruta / total_ingresos * 100, 2) if total_ingresos else 0
        ganancia_antes_gastos_fin = ganancia_bruta - total_gastos_op
        pct_antes_gastos_fin    = round(ganancia_antes_gastos_fin / total_ingresos * 100, 2) if total_ingresos else 0
        ganancia_antes_imp      = ganancia_antes_gastos_fin - total_gastos_fin - total_otros
        ganancia_neta           = ganancia_antes_imp - total_impuestos
        pct_neta                = round(ganancia_neta / total_ingresos * 100, 2) if total_ingresos else 0

    return jsonify({
        'ingresos': ingresos, 'total_ingresos': total_ingresos,
        'costos': costos, 'total_costos': total_costos,
        'gastos_operativos': gastos_op, 'total_gastos_op': total_gastos_op,
        'gastos_financieros': gastos_fin, 'total_gastos_fin': total_gastos_fin,
        'otros_gastos': otros_gastos, 'total_otros': total_otros,
        'impuestos': impuestos_list, 'total_impuestos': total_impuestos,
        'ganancia_bruta': ganancia_bruta, 'pct_bruta': pct_bruta,
        'ganancia_antes_gastos_fin': ganancia_antes_gastos_fin, 'pct_antes_gastos_fin': pct_antes_gastos_fin,
        'ganancia_antes_impuestos': ganancia_antes_imp,
        'ganancia_neta': ganancia_neta, 'pct_neta': pct_neta,
        # backward compat for dashboard
        'total_gastos_admin': total_gastos_fin,
        'utilidad_bruta': ganancia_bruta,
        'utilidad_operativa': ganancia_antes_gastos_fin,
        'utilidad_neta': ganancia_neta,
    })

# ─── PÉRDIDAS ────────────────────────────────────────
@app.route('/api/losses', methods=['GET'])
def get_losses():
    with get_db() as conn:
        rows = conn.execute('SELECT * FROM losses ORDER BY created_at DESC').fetchall()
    return jsonify(rows_to_list(rows))

@app.route('/api/losses', methods=['POST'])
def create_loss():
    d = request.json or {}
    if not d.get('product_name') or not d.get('quantity') or not d.get('reason') or not d.get('responsible'):
        return jsonify({'error': 'Producto, cantidad, motivo y responsable son requeridos'}), 400
    with get_db() as conn:
        cur = conn.execute(
            'INSERT INTO losses (product_id, product_name, quantity, unit, reason, responsible, notes, sale_value, category) VALUES (?,?,?,?,?,?,?,?,?)',
            (d.get('product_id'), d['product_name'], d['quantity'],
             d.get('unit', 'unidad'), d['reason'], d['responsible'], d.get('notes', ''),
             d.get('sale_value', 0), d.get('category', ''))
        )
        # Descontar del inventario si el producto tiene id
        if d.get('product_id'):
            conn.execute(
                'UPDATE products SET stock = MAX(0, stock - ?) WHERE id=?',
                (d['quantity'], d['product_id'])
            )
        row = conn.execute('SELECT * FROM losses WHERE id=?', (cur.lastrowid,)).fetchone()
    return jsonify(row_to_dict(row))

@app.route('/api/losses/<int:lid>', methods=['DELETE'])
def delete_loss(lid):
    with get_db() as conn:
        loss = conn.execute('SELECT * FROM losses WHERE id=?', (lid,)).fetchone()
        if not loss:
            return jsonify({'error': 'Pérdida no encontrada'}), 404
        # Restaurar stock
        if loss['product_id']:
            conn.execute(
                'UPDATE products SET stock = stock + ? WHERE id=?',
                (loss['quantity'], loss['product_id'])
            )
        conn.execute('DELETE FROM losses WHERE id=?', (lid,))
    return jsonify({'ok': True})

# ─── TERMINALES POS ──────────────────────────────────
@app.route('/api/terminals', methods=['GET'])
def get_terminals():
    with get_db() as conn:
        rows = conn.execute('SELECT * FROM pos_terminals WHERE active=1 ORDER BY id').fetchall()
    return jsonify(rows_to_list(rows))

@app.route('/api/terminals', methods=['POST'])
def create_terminal():
    d = request.json or {}
    if not d.get('name'):
        return jsonify({'error': 'Nombre requerido'}), 400
    with get_db() as conn:
        cur = conn.execute('INSERT INTO pos_terminals (name, color, icon) VALUES (?,?,?)',
                           (d['name'], d.get('color','#3b82f6'), d.get('icon','🛒')))
        row = conn.execute('SELECT * FROM pos_terminals WHERE id=?', (cur.lastrowid,)).fetchone()
    return jsonify(row_to_dict(row))

@app.route('/api/terminals/<int:tid>', methods=['DELETE'])
def delete_terminal(tid):
    with get_db() as conn:
        conn.execute('UPDATE pos_terminals SET active=0 WHERE id=?', (tid,))
    return jsonify({'ok': True})

# ─── SETTINGS ────────────────────────────────────────
@app.route('/api/settings', methods=['GET'])
def get_settings():
    with get_db() as conn:
        rows = conn.execute('SELECT key, value FROM settings').fetchall()
    return jsonify({r['key']: r['value'] for r in rows})

@app.route('/api/settings', methods=['POST'])
def save_settings():
    d = request.json or {}
    with get_db() as conn:
        for key, value in d.items():
            conn.execute('INSERT OR REPLACE INTO settings (key, value) VALUES (?,?)', (key, str(value)))
    return jsonify({'ok': True})

@app.route('/api/settings/test-whatsapp', methods=['POST'])
def test_whatsapp():
    d = request.json or {}
    phone    = d.get('phone', '').strip()
    instance = d.get('greenapi_instance', '').strip()
    token    = d.get('greenapi_token', '').strip()
    if not phone or not instance or not token:
        return jsonify({'error': 'Completa el número, ID de instancia y token'}), 400
    try:
        import json as _json
        chat_id = phone.lstrip('+').replace(' ', '') + '@c.us'
        url = f'https://api.green-api.com/waInstance{instance}/sendMessage/{token}'
        payload = _json.dumps({'chatId': chat_id, 'message': '✅ Prueba exitosa — POS Delicias del Rey conectado.'}).encode()
        req = urllib.request.Request(url, data=payload, headers={'Content-Type': 'application/json'})
        resp = urllib.request.urlopen(req, timeout=15)
        result = _json.loads(resp.read().decode())
        print(f'[GreenAPI] Respuesta: {result}')
        return jsonify({'ok': True, 'result': result})
    except urllib.error.HTTPError as e:
        body = e.read().decode()
        print(f'[GreenAPI] HTTP Error {e.code}: {body}')
        return jsonify({'error': f'Error {e.code}: {body}'}), 400
    except Exception as e:
        print(f'[GreenAPI] Error: {e}')
        return jsonify({'error': str(e)}), 500

# ─── TRABAJADORES ────────────────────────────────────
@app.route('/api/workers', methods=['GET'])
def get_workers():
    with get_db() as conn:
        rows = conn.execute('SELECT * FROM workers WHERE active=1 ORDER BY name').fetchall()
    return jsonify(rows_to_list(rows))

@app.route('/api/workers', methods=['POST'])
def create_worker():
    d = request.json or {}
    if not d.get('name'):
        return jsonify({'error': 'El nombre es requerido'}), 400
    with get_db() as conn:
        cur = conn.execute(
            'INSERT INTO workers (name, cargo, phone) VALUES (?,?,?)',
            (d['name'], d.get('cargo',''), d.get('phone',''))
        )
        row = conn.execute('SELECT * FROM workers WHERE id=?', (cur.lastrowid,)).fetchone()
    return jsonify(row_to_dict(row))

@app.route('/api/workers/<int:wid>', methods=['DELETE'])
def delete_worker(wid):
    with get_db() as conn:
        conn.execute('UPDATE workers SET active=0 WHERE id=?', (wid,))
    return jsonify({'ok': True})

@app.route('/api/workers/<int:wid>/notes', methods=['GET'])
def get_worker_notes(wid):
    with get_db() as conn:
        rows = conn.execute(
            'SELECT * FROM worker_notes WHERE worker_id=? ORDER BY created_at DESC', (wid,)
        ).fetchall()
    return jsonify(rows_to_list(rows))

@app.route('/api/workers/<int:wid>/notes', methods=['POST'])
def create_worker_note(wid):
    d = request.json or {}
    if not d.get('content'):
        return jsonify({'error': 'El contenido es requerido'}), 400
    with get_db() as conn:
        worker = conn.execute('SELECT id FROM workers WHERE id=? AND active=1', (wid,)).fetchone()
        if not worker:
            return jsonify({'error': 'Trabajador no encontrado'}), 404
        note_type = d.get('note_type', 'nota')
        cur = conn.execute(
            'INSERT INTO worker_notes (worker_id, content, note_type) VALUES (?,?,?)',
            (wid, d['content'], note_type)
        )
        row = conn.execute('SELECT * FROM worker_notes WHERE id=?', (cur.lastrowid,)).fetchone()
    return jsonify(row_to_dict(row))

@app.route('/api/workers/notes/<int:nid>', methods=['DELETE'])
def delete_worker_note(nid):
    with get_db() as conn:
        conn.execute('DELETE FROM worker_notes WHERE id=?', (nid,))
    return jsonify({'ok': True})

@app.route('/api/workers/<int:wid>/stats', methods=['GET'])
def get_worker_stats(wid):
    with get_db() as conn:
        worker = conn.execute('SELECT * FROM workers WHERE id=?', (wid,)).fetchone()
        if not worker:
            return jsonify({'error': 'Not found'}), 404
        name = worker['name']
        # Descuadres (cierres con diferencia negativa donde fue responsable)
        desc = conn.execute(
            "SELECT COUNT(*) as cnt, COALESCE(SUM(ABS(difference)),0) as total FROM cash_registers WHERE status='cerrada' AND notes LIKE ? AND difference < 0",
            (f'%{name}%',)
        ).fetchone()
        # Pérdidas registradas
        perd = conn.execute(
            "SELECT COUNT(*) as cnt, COALESCE(SUM(sale_value),0) as valor FROM losses WHERE responsible LIKE ?",
            (f'%{name}%',)
        ).fetchone()
        # Tareas pendientes
        tareas_pend = conn.execute(
            "SELECT COUNT(*) as cnt FROM tasks WHERE assigned_to LIKE ? AND status='pendiente'",
            (f'%{name}%',)
        ).fetchone()
        tareas_total = conn.execute(
            "SELECT COUNT(*) as cnt FROM tasks WHERE assigned_to LIKE ?",
            (f'%{name}%',)
        ).fetchone()
        # Notas
        notas_cnt = conn.execute(
            "SELECT COUNT(*) as cnt FROM worker_notes WHERE worker_id=?", (wid,)
        ).fetchone()
        # Últimos descuadres
        ultimos = conn.execute(
            "SELECT opened_at, closed_at, difference, notes FROM cash_registers WHERE status='cerrada' AND notes LIKE ? AND difference < 0 ORDER BY closed_at DESC LIMIT 5",
            (f'%{name}%',)
        ).fetchall()
        # Lista de pérdidas del trabajador
        perdidas_list = conn.execute(
            "SELECT product_name, quantity, unit, reason, category, sale_value, created_at FROM losses WHERE responsible LIKE ? ORDER BY created_at DESC",
            (f'%{name}%',)
        ).fetchall()
    return jsonify({
        'descuadres': {'count': desc['cnt'], 'total': desc['total']},
        'perdidas': {'count': perd['cnt'], 'valor': perd['valor']},
        'tareas': {'pendientes': tareas_pend['cnt'], 'total': tareas_total['cnt']},
        'notas': notas_cnt['cnt'],
        'ultimos_descuadres': rows_to_list(ultimos),
        'perdidas_list': rows_to_list(perdidas_list)
    })

@app.route('/api/workers/<int:wid>/documents', methods=['GET'])
def get_worker_documents(wid):
    with get_db() as conn:
        rows = conn.execute(
            'SELECT * FROM worker_documents WHERE worker_id=? ORDER BY created_at DESC', (wid,)
        ).fetchall()
    return jsonify(rows_to_list(rows))

@app.route('/api/workers/<int:wid>/documents', methods=['POST'])
def upload_worker_document(wid):
    import time as _time
    doc_type    = request.form.get('doc_type', 'general')
    description = request.form.get('description', '')
    filename    = ''
    original_name = ''
    upload_dir  = os.path.join(os.path.dirname(__file__), 'public', 'uploads', 'workers')
    os.makedirs(upload_dir, exist_ok=True)
    if 'file' in request.files:
        f = request.files['file']
        if f and f.filename:
            ext = f.filename.rsplit('.', 1)[-1].lower() if '.' in f.filename else 'pdf'
            filename = f'worker_{wid}_{int(_time.time())}.{ext}'
            original_name = f.filename
            f.save(os.path.join(upload_dir, filename))
    with get_db() as conn:
        worker = conn.execute('SELECT id FROM workers WHERE id=? AND active=1', (wid,)).fetchone()
        if not worker:
            return jsonify({'error': 'Trabajador no encontrado'}), 404
        cur = conn.execute(
            'INSERT INTO worker_documents (worker_id, filename, original_name, doc_type, description) VALUES (?,?,?,?,?)',
            (wid, filename, original_name, doc_type, description)
        )
        row = conn.execute('SELECT * FROM worker_documents WHERE id=?', (cur.lastrowid,)).fetchone()
    return jsonify(row_to_dict(row))

@app.route('/api/workers/documents/<int:did>', methods=['DELETE'])
def delete_worker_document(did):
    with get_db() as conn:
        doc = conn.execute('SELECT * FROM worker_documents WHERE id=?', (did,)).fetchone()
        if doc and doc['filename']:
            path = os.path.join(os.path.dirname(__file__), 'public', 'uploads', 'workers', doc['filename'])
            if os.path.exists(path):
                os.remove(path)
        conn.execute('DELETE FROM worker_documents WHERE id=?', (did,))
    return jsonify({'ok': True})

# ─── CHECKLIST ───────────────────────────────────────
@app.route('/api/checklist', methods=['GET'])
def get_checklist():
    turno = request.args.get('turno', 'apertura')
    date  = request.args.get('date', datetime.now().strftime('%Y-%m-%d'))
    with get_db() as conn:
        items = conn.execute(
            'SELECT * FROM checklist_items WHERE turno=? AND active=1 ORDER BY order_num',
            (turno,)
        ).fetchall()
        logs = conn.execute(
            'SELECT item_id, completed_by, created_at FROM checklist_logs WHERE date=? AND turno=?',
            (date, turno)
        ).fetchall()
        done_map = {r['item_id']: {'by': r['completed_by'], 'at': r['created_at']} for r in logs}
        result = []
        for it in items:
            d = dict(it)
            d['completed'] = it['id'] in done_map
            d['completed_by'] = done_map.get(it['id'], {}).get('by', '')
            d['completed_at'] = done_map.get(it['id'], {}).get('at', '')
            result.append(d)
    return jsonify(result)

@app.route('/api/checklist/toggle', methods=['POST'])
def toggle_checklist():
    d = request.json or {}
    item_id      = d.get('item_id')
    date         = d.get('date', datetime.now().strftime('%Y-%m-%d'))
    turno        = d.get('turno', 'apertura')
    completed    = d.get('completed', True)
    completed_by = d.get('completed_by', '')
    with get_db() as conn:
        if completed:
            conn.execute(
                'INSERT OR REPLACE INTO checklist_logs (date, turno, item_id, completed_by) VALUES (?,?,?,?)',
                (date, turno, item_id, completed_by)
            )
        else:
            conn.execute(
                'DELETE FROM checklist_logs WHERE date=? AND turno=? AND item_id=?',
                (date, turno, item_id)
            )
    return jsonify({'ok': True})

@app.route('/api/checklist/history', methods=['GET'])
def get_checklist_history():
    turno = request.args.get('turno', 'apertura')
    with get_db() as conn:
        rows = conn.execute("""
            SELECT l.date, l.turno,
                   COUNT(*) as completados,
                   (SELECT COUNT(*) FROM checklist_items WHERE turno=l.turno AND active=1) as total
            FROM checklist_logs l
            WHERE l.turno=?
            GROUP BY l.date, l.turno
            ORDER BY l.date DESC
            LIMIT 14
        """, (turno,)).fetchall()
    return jsonify(rows_to_list(rows))

@app.route('/api/aseo', methods=['GET'])
def get_aseo():
    dia  = request.args.get('dia', 'lunes')
    date = request.args.get('date', datetime.now().strftime('%Y-%m-%d'))
    with get_db() as conn:
        items = conn.execute(
            'SELECT * FROM aseo_items WHERE dia=? AND active=1 ORDER BY order_num',
            (dia,)
        ).fetchall()
        logs = conn.execute(
            'SELECT item_id, completed_by, created_at FROM aseo_logs WHERE date=? AND dia=?',
            (date, dia)
        ).fetchall()
        done_map = {r['item_id']: {'by': r['completed_by'], 'at': r['created_at']} for r in logs}
        result = []
        for it in items:
            d = dict(it)
            d['completed']    = it['id'] in done_map
            d['completed_by'] = done_map.get(it['id'], {}).get('by', '')
            d['completed_at'] = done_map.get(it['id'], {}).get('at', '')
            result.append(d)
    return jsonify(result)

@app.route('/api/aseo/toggle', methods=['POST'])
def toggle_aseo():
    d            = request.json or {}
    item_id      = d.get('item_id')
    date         = d.get('date', datetime.now().strftime('%Y-%m-%d'))
    dia          = d.get('dia', 'lunes')
    completed    = d.get('completed', True)
    completed_by = d.get('completed_by', '')
    with get_db() as conn:
        if completed:
            conn.execute(
                'INSERT OR REPLACE INTO aseo_logs (date, dia, item_id, completed_by) VALUES (?,?,?,?)',
                (date, dia, item_id, completed_by)
            )
        else:
            conn.execute(
                'DELETE FROM aseo_logs WHERE date=? AND dia=? AND item_id=?',
                (date, dia, item_id)
            )
    return jsonify({'ok': True})

@app.route('/api/aseo/items', methods=['POST'])
def add_aseo_item():
    d    = request.json or {}
    dia  = d.get('dia', 'lunes')
    text = d.get('text', '').strip()
    if not text:
        return jsonify({'error': 'text requerido'}), 400
    with get_db() as conn:
        max_order = conn.execute(
            'SELECT COALESCE(MAX(order_num),0) FROM aseo_items WHERE dia=?', (dia,)
        ).fetchone()[0]
        cur = conn.execute(
            'INSERT INTO aseo_items (dia, text, order_num) VALUES (?,?,?)',
            (dia, text, max_order + 1)
        )
    return jsonify({'ok': True, 'id': cur.lastrowid})

@app.route('/api/aseo/items/<int:item_id>', methods=['DELETE'])
def delete_aseo_item(item_id):
    with get_db() as conn:
        conn.execute('UPDATE aseo_items SET active=0 WHERE id=?', (item_id,))
    return jsonify({'ok': True})

# ─── NOVEDADES ──────────────────────────────────────────
@app.route('/api/novedades', methods=['GET'])
def get_novedades():
    solo_nuevas = request.args.get('nuevas', '0') == '1'
    with get_db() as conn:
        if solo_nuevas:
            rows = conn.execute(
                'SELECT * FROM novedades WHERE visto=0 ORDER BY created_at DESC'
            ).fetchall()
        else:
            rows = conn.execute(
                'SELECT * FROM novedades ORDER BY created_at DESC LIMIT 50'
            ).fetchall()
    return jsonify(rows_to_list(rows))

@app.route('/api/novedades', methods=['POST'])
def create_novedad():
    d            = request.json or {}
    tipo         = d.get('tipo', 'otro')
    descripcion  = d.get('descripcion', '').strip()
    reportado_por = d.get('reportado_por', '')
    if not descripcion:
        return jsonify({'error': 'descripcion requerida'}), 400
    with get_db() as conn:
        cur = conn.execute(
            'INSERT INTO novedades (tipo, descripcion, reportado_por) VALUES (?,?,?)',
            (tipo, descripcion, reportado_por)
        )
    return jsonify({'ok': True, 'id': cur.lastrowid})

@app.route('/api/novedades/marcar-vistas', methods=['POST'])
def marcar_novedades_vistas():
    with get_db() as conn:
        conn.execute('UPDATE novedades SET visto=1 WHERE visto=0')
    return jsonify({'ok': True})

@app.route('/api/novedades/<int:nid>', methods=['DELETE'])
def delete_novedad(nid):
    with get_db() as conn:
        conn.execute('DELETE FROM novedades WHERE id=?', (nid,))
    return jsonify({'ok': True})

# ─── CUENTAS ABIERTAS ──────────────────────────────────
@app.route('/api/open-accounts', methods=['GET'])
def get_open_accounts():
    with get_db() as conn:
        rows = conn.execute(
            'SELECT * FROM open_accounts ORDER BY created_at ASC'
        ).fetchall()
    result = []
    for r in rows:
        d = dict(r)
        try:
            d['items'] = json.loads(d.get('items') or '[]')
        except Exception:
            d['items'] = []
        result.append(d)
    return jsonify(result)

@app.route('/api/open-accounts', methods=['POST'])
def create_open_account():
    d = request.json or {}
    name = (d.get('name') or 'Cuenta').strip() or 'Cuenta'
    items = d.get('items', [])
    with get_db() as conn:
        cur = conn.execute(
            'INSERT INTO open_accounts (name, items) VALUES (?,?)',
            (name, json.dumps(items))
        )
        row = conn.execute('SELECT * FROM open_accounts WHERE id=?', (cur.lastrowid,)).fetchone()
    res = dict(row)
    res['items'] = items
    return jsonify(res)

@app.route('/api/open-accounts/<int:aid>', methods=['PUT'])
def update_open_account(aid):
    d = request.json or {}
    with get_db() as conn:
        row = conn.execute('SELECT * FROM open_accounts WHERE id=?', (aid,)).fetchone()
        if not row:
            return jsonify({'error': 'Cuenta no encontrada'}), 404
        name  = (d.get('name') if d.get('name') is not None else row['name'])
        items = d.get('items') if d.get('items') is not None else json.loads(row['items'] or '[]')
        conn.execute(
            "UPDATE open_accounts SET name=?, items=?, updated_at=datetime('now','localtime') WHERE id=?",
            (name, json.dumps(items), aid)
        )
    return jsonify({'ok': True})

@app.route('/api/open-accounts/<int:aid>', methods=['DELETE'])
def delete_open_account(aid):
    with get_db() as conn:
        conn.execute('DELETE FROM open_accounts WHERE id=?', (aid,))
    return jsonify({'ok': True})

# ─── RECEPCIÓN DE INSUMOS ──────────────────────────────
@app.route('/api/recepciones', methods=['GET'])
def get_recepciones():
    estado = request.args.get('estado')  # 'pendiente' | 'recibido' | None
    with get_db() as conn:
        if estado in ('pendiente', 'recibido'):
            rows = conn.execute(
                'SELECT * FROM recepciones WHERE estado=? ORDER BY '
                'CASE WHEN estado="pendiente" THEN fecha_esperada END ASC, '
                'recibido_at DESC, created_at DESC',
                (estado,)
            ).fetchall()
        else:
            rows = conn.execute(
                'SELECT * FROM recepciones ORDER BY created_at DESC LIMIT 200'
            ).fetchall()
    return jsonify(rows_to_list(rows))

@app.route('/api/recepciones', methods=['POST'])
def create_recepcion():
    # Acepta multipart (con foto) o JSON
    if request.content_type and 'multipart/form-data' in request.content_type:
        f = request.form
        descripcion    = (f.get('descripcion') or '').strip()
        cantidad       = (f.get('cantidad') or '').strip()
        proveedor      = (f.get('proveedor') or '').strip()
        fecha_esperada = (f.get('fecha_esperada') or '').strip()
        created_by     = (f.get('created_by') or '').strip()
        foto_file      = request.files.get('foto')
    else:
        d = request.json or {}
        descripcion    = (d.get('descripcion') or '').strip()
        cantidad       = (d.get('cantidad') or '').strip()
        proveedor      = (d.get('proveedor') or '').strip()
        fecha_esperada = (d.get('fecha_esperada') or '').strip()
        created_by     = (d.get('created_by') or '').strip()
        foto_file      = None
    if not descripcion:
        return jsonify({'error': 'descripcion requerida'}), 400
    with get_db() as conn:
        cur = conn.execute(
            'INSERT INTO recepciones (descripcion, cantidad, proveedor, fecha_esperada, created_by) '
            'VALUES (?,?,?,?,?)',
            (descripcion, cantidad, proveedor, fecha_esperada, created_by)
        )
        rid = cur.lastrowid
        foto_name = ''
        if foto_file and foto_file.filename:
            ext = foto_file.filename.rsplit('.', 1)[-1].lower()
            if ext in ('jpg', 'jpeg', 'png', 'gif', 'webp'):
                upload_dir = os.path.join(os.path.dirname(__file__), 'public', 'uploads', 'recepciones')
                os.makedirs(upload_dir, exist_ok=True)
                foto_name = f'recepcion_{rid}.{ext}'
                foto_file.save(os.path.join(upload_dir, foto_name))
                conn.execute('UPDATE recepciones SET foto=? WHERE id=?', (foto_name, rid))
        row = conn.execute('SELECT * FROM recepciones WHERE id=?', (rid,)).fetchone()
    return jsonify(row_to_dict(row))

@app.route('/api/recepciones/<int:rid>/recibir', methods=['POST'])
def recibir_recepcion(rid):
    d = request.json or {}
    recibido_por = (d.get('recibido_por') or '').strip()
    nota         = (d.get('nota') or '').strip()
    recibido     = d.get('recibido', True)
    with get_db() as conn:
        if recibido:
            conn.execute(
                "UPDATE recepciones SET estado='recibido', recibido_por=?, nota_recepcion=?, "
                "recibido_at=datetime('now','localtime') WHERE id=?",
                (recibido_por, nota, rid)
            )
        else:
            conn.execute(
                "UPDATE recepciones SET estado='pendiente', recibido_por='', nota_recepcion='', "
                "recibido_at=NULL WHERE id=?",
                (rid,)
            )
        row = conn.execute('SELECT * FROM recepciones WHERE id=?', (rid,)).fetchone()
    return jsonify(row_to_dict(row))

@app.route('/api/recepciones/<int:rid>', methods=['DELETE'])
def delete_recepcion(rid):
    with get_db() as conn:
        row = conn.execute('SELECT foto FROM recepciones WHERE id=?', (rid,)).fetchone()
        if row and row['foto']:
            fp = os.path.join(os.path.dirname(__file__), 'public', 'uploads', 'recepciones', row['foto'])
            if os.path.exists(fp):
                try: os.remove(fp)
                except Exception: pass
        conn.execute('DELETE FROM recepciones WHERE id=?', (rid,))
    return jsonify({'ok': True})

@app.route('/api/checklist/items', methods=['POST'])
def add_checklist_item():
    d = request.json or {}
    turno    = d.get('turno', 'apertura')
    section  = d.get('section', '').strip()
    text     = d.get('text', '').strip()
    if not text:
        return jsonify({'error': 'text requerido'}), 400
    with get_db() as conn:
        max_order = conn.execute(
            'SELECT COALESCE(MAX(order_num),0) FROM checklist_items WHERE turno=?', (turno,)
        ).fetchone()[0]
        cur = conn.execute(
            'INSERT INTO checklist_items (turno, section, text, order_num) VALUES (?,?,?,?)',
            (turno, section, text, max_order + 1)
        )
    return jsonify({'ok': True, 'id': cur.lastrowid})

@app.route('/api/checklist/items/<int:item_id>', methods=['DELETE'])
def delete_checklist_item(item_id):
    with get_db() as conn:
        conn.execute('UPDATE checklist_items SET active=0 WHERE id=?', (item_id,))
    return jsonify({'ok': True})

# ─── TAREAS ──────────────────────────────────────────
@app.route('/api/tasks', methods=['GET'])
def get_tasks():
    with get_db() as conn:
        rows = conn.execute('SELECT * FROM tasks ORDER BY status ASC, created_at DESC').fetchall()
    return jsonify(rows_to_list(rows))

@app.route('/api/tasks', methods=['POST'])
def create_task():
    d = request.json or {}
    if not d.get('funcion'):
        return jsonify({'error': 'La función es requerida'}), 400
    with get_db() as conn:
        cur = conn.execute(
            'INSERT INTO tasks (funcion, area, assigned_to) VALUES (?,?,?)',
            (d['funcion'], d.get('area',''), d.get('assigned_to',''))
        )
        row = conn.execute('SELECT * FROM tasks WHERE id=?', (cur.lastrowid,)).fetchone()
    return jsonify(row_to_dict(row))

@app.route('/api/tasks/<int:tid>/status', methods=['PUT'])
def update_task_status(tid):
    d = request.json or {}
    new_status = d.get('status')
    if new_status not in ('pendiente', 'realizado'):
        return jsonify({'error': 'Estado inválido'}), 400
    completed_at = datetime.now().strftime('%Y-%m-%d %H:%M:%S') if new_status == 'realizado' else None
    with get_db() as conn:
        conn.execute('UPDATE tasks SET status=?, completed_at=? WHERE id=?', (new_status, completed_at, tid))
        row = conn.execute('SELECT * FROM tasks WHERE id=?', (tid,)).fetchone()
    return jsonify(row_to_dict(row))

@app.route('/api/tasks/<int:tid>', methods=['DELETE'])
def delete_task(tid):
    with get_db() as conn:
        conn.execute('DELETE FROM tasks WHERE id=?', (tid,))
    return jsonify({'ok': True})

# ─── RESET ───────────────────────────────────────────
@app.route('/api/reset', methods=['POST'])
def reset_pos():
    d = request.json or {}
    pin = str(d.get('pin', '')).strip()
    if not pin:
        return jsonify({'error': 'PIN requerido'}), 400

    with get_db() as conn:
        cfg = {r['key']: r['value'] for r in conn.execute('SELECT key,value FROM settings').fetchall()}

    pin_admin  = cfg.get('pin_admin',  '1623')
    pin_worker = cfg.get('pin_worker', '0000')

    if pin == pin_admin:
        # Borrar TODO
        with get_db() as conn:
            conn.executescript("""
                DELETE FROM sale_items;
                DELETE FROM sales;
                DELETE FROM cash_movements;
                DELETE FROM cash_registers;
                DELETE FROM comandas;
                DELETE FROM return_items;
                DELETE FROM returns;
                DELETE FROM journal_entries;
                DELETE FROM losses;
                DELETE FROM products;
                UPDATE SQLITE_SEQUENCE SET seq=0 WHERE name IN
                    ('sales','sale_items','cash_movements','cash_registers',
                     'comandas','returns','return_items','journal_entries',
                     'losses','products');
            """)
            # Reinsertar productos de muestra
            samples = [
                ('Café americano','Bebidas',25,100,'taza',10),
                ('Café con leche','Bebidas',30,100,'taza',10),
                ('Capuchino','Bebidas',35,100,'taza',10),
                ('Té negro','Bebidas',20,50,'taza',10),
                ('Agua natural 500ml','Bebidas',15,24,'botella',6),
                ('Croissant','Panadería',28,20,'pieza',5),
                ('Muffin','Panadería',25,15,'pieza',5),
                ('Sándwich jamón','Comida',55,10,'pieza',3),
                ('Ensalada de frutas','Comida',45,8,'porción',3),
                ('Jugo naranja','Bebidas',35,10,'vaso',5),
            ]
            conn.executemany(
                'INSERT INTO products (name,category,price,stock,unit,low_stock_alert) VALUES (?,?,?,?,?,?)',
                samples
            )
        return jsonify({'ok': True, 'type': 'full'})

    elif pin == pin_worker:
        # Borrar solo inventario
        with get_db() as conn:
            conn.executescript("""
                DELETE FROM products;
                UPDATE SQLITE_SEQUENCE SET seq=0 WHERE name='products';
            """)
            samples = [
                ('Café americano','Bebidas',25,100,'taza',10),
                ('Café con leche','Bebidas',30,100,'taza',10),
                ('Capuchino','Bebidas',35,100,'taza',10),
                ('Té negro','Bebidas',20,50,'taza',10),
                ('Agua natural 500ml','Bebidas',15,24,'botella',6),
                ('Croissant','Panadería',28,20,'pieza',5),
                ('Muffin','Panadería',25,15,'pieza',5),
                ('Sándwich jamón','Comida',55,10,'pieza',3),
                ('Ensalada de frutas','Comida',45,8,'porción',3),
                ('Jugo naranja','Bebidas',35,10,'vaso',5),
            ]
            conn.executemany(
                'INSERT INTO products (name,category,price,stock,unit,low_stock_alert) VALUES (?,?,?,?,?,?)',
                samples
            )
        return jsonify({'ok': True, 'type': 'inventory'})

    else:
        return jsonify({'error': 'PIN incorrecto'}), 401


def send_whatsapp(phone, instance, token, message):
    try:
        import json as _json
        # Formatear chatId: número colombiano → 57XXXXXXXXXX@c.us
        chat_id = phone.lstrip('+').replace(' ', '') + '@c.us'
        url = f'https://api.green-api.com/waInstance{instance}/sendMessage/{token}'
        payload = _json.dumps({'chatId': chat_id, 'message': message}).encode()
        req = urllib.request.Request(url, data=payload, headers={'Content-Type': 'application/json'})
        urllib.request.urlopen(req, timeout=15)
        print(f'[WhatsApp GreenAPI] Mensaje enviado a {chat_id}')
    except Exception as e:
        print(f'[WhatsApp GreenAPI] Error: {e}')

# ─── Inicialización al arrancar (también bajo Gunicorn/producción) ───
# Se ejecuta al importar el módulo, no solo con `python server.py`,
# para que las tablas e índices existan cuando corre con Gunicorn.
init_db()
with get_db() as conn:
    conn.executescript("""
        CREATE INDEX IF NOT EXISTS idx_sales_created ON sales(created_at);
        CREATE INDEX IF NOT EXISTS idx_sale_items_sale ON sale_items(sale_id);
        CREATE INDEX IF NOT EXISTS idx_movements_created ON cash_movements(created_at);
        CREATE INDEX IF NOT EXISTS idx_products_active ON products(active, category);
        CREATE INDEX IF NOT EXISTS idx_comandas_status ON comandas(status);
    """)

if __name__ == '__main__':
    print('\n✅ POS Cafetería corriendo en http://localhost:3000\n')
    port = int(os.environ.get('PORT', 3000))
    app.run(host='0.0.0.0', port=port, debug=False, threaded=True)
